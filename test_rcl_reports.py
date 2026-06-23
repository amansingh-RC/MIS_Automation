"""Self-test for the Scrap + Stock combined report."""

import io

import openpyxl

from rcl_reports import process_scrap_and_stock, _r3, process_scrap, process_stock

# A raw export with BOTH the columns the two tools need.
HEADERS = ["Wcgroup Name", "Wc Name", "Party Name", "Item Group",
           "Variantt Name", "Gross Weight", "Metal Weight", "Stock Status"]

# wcgroup, wcname, party, item, variant, gross, metal, status
ROWS = [
    # --- scrap rows (Stock Status SCRAP / HL-SCRAP) ---
    ["GRP-A", "WC-1", "", "GOLD", "", 10.0, 8.0, "SCRAP"],
    ["GRP-A", "WC-1", "", "GOLD", "", 5.0, 4.0, "HL-SCRAP"],   # merges with above
    ["GRP-A", "WC-2", "", "GOLD", "", 2.0, 1.5, "SCRAP"],
    ["GRP-B", "WC-9", "", "GOLD", "", 7.0, 6.0, "scrap"],      # lowercase ok
    ["GRP-A", "WC-3", "", "GOLD", "", 1.0, 1.0, "STOCK"],      # not scrap -> excluded
    # --- stock-only behavior rows ---
    ["", "", "PARTY-X", "GOLD", "", 3.0, 2.0, "STOCK"],        # blank WC -> party fill
    ["GRP-A", "WC-1", "", "ALLOY", "", 99.0, 99.0, "STOCK"],   # ALLOY removed (stock)
    ["GRP-C", "WC-5", "", "OTHER METAL", "IRON", 50.0, 50.0, "STOCK"],  # non-OM removed
    ["GRP-C", "WC-5", "", "OTHER METAL", "OM", 4.0, 3.0, "STOCK"],      # OM kept
    ["Grand Total", "", "", "GOLD", "", 999.0, 999.0, "STOCK"],         # total row skipped
]


def make_input() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Some Title"])           # noise row
    ws.append(HEADERS)
    for r in ROWS:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def main():
    data = make_input()

    # Exercise the individual processors for precise assertions.
    from rcl_reports import _grid_first_sheet
    grid = _grid_first_sheet(data)

    # --- SCRAP ---
    scrap = process_scrap(grid)
    # scrap rows: 4 (two GRP-A/WC-1 merge, GRP-A/WC-2, GRP-B/WC-9)
    assert scrap["row_count"] == 4, scrap["row_count"]
    ga = scrap["groups"]["GRP-A"]
    wc1 = next(e for e in ga if e["wcname"] == "WC-1")
    assert _r3(wc1["gross"]) == 15.0 and _r3(wc1["metal"]) == 12.0, wc1
    assert _r3(scrap["total_gross"]) == 24.0, scrap["total_gross"]   # 10+5+2+7
    print("OK  scrap grouping + merge + total")

    # --- STOCK (does NOT filter by Stock Status) ---
    stock = process_stock(grid)
    # Included: GRP-A WC-1 (10+5), WC-2 (2), WC-3 (1); GRP-B WC-9 (7);
    # PARTY-X (3, party-filled); GRP-C WC-5 OM (4).
    # Excluded: ALLOY row, non-OM OTHER METAL row, Grand Total row.
    assert stock["filtered_rows"] == 7, stock["filtered_rows"]
    assert _r3(stock["total_gross"]) == 32.0, stock["total_gross"]
    groups = stock["groups"]
    assert "PARTY-X" in groups, list(groups)        # blank WC filled from party
    assert "GRP-C" in groups
    assert "GRP-A" in groups
    # GRP-A/WC-1 must be 15 (ALLOY row excluded, not added)
    wc1_stock = next(e for e in groups["GRP-A"] if e["wcname"] == "WC-1")
    assert _r3(wc1_stock["gross"]) == 15.0, wc1_stock
    # GRP-C/WC-5 must be 4 (only OM kept)
    wc5 = groups["GRP-C"][0]
    assert _r3(wc5["gross"]) == 4.0, wc5
    print("OK  stock filters (ALLOY/OM/total/party-fill)")

    # --- combined workbook ---
    out_bytes, summary = process_scrap_and_stock(data, "2026-06-23")
    wb = openpyxl.load_workbook(io.BytesIO(out_bytes))
    assert wb.sheetnames == ["SCRAP REPORT", "STOCK REPORT"], wb.sheetnames
    s1 = wb["SCRAP REPORT"]
    assert s1["A1"].value == "ROYAL CHAIN LIMITED"
    assert s1["A4"].value == "Wcgroup Name"
    # find Grand Total row in scrap sheet
    grand = [r for r in s1.iter_rows(values_only=True)
             if r and r[0] == "Grand Total"][0]
    assert _r3(grand[2]) == 24.0, grand
    s2 = wb["STOCK REPORT"]
    assert s2["A5"].value == "Wcgroup Name"          # header on row 5 (3 title rows + blank)
    print("OK  combined workbook: 2 sheets, structure, grand totals")

    assert summary["scrap_rows"] == 4
    assert summary["stock_filtered"] == 7
    print("OK  summary stats")
    print("\nALL TESTS PASSED")


if __name__ == "__main__":
    main()
