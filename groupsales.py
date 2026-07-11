from __future__ import annotations

import re
from collections import OrderedDict
from dataclasses import dataclass

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.colors import Color

from loss_report import LossReportError, _grid_from_bytes, to_number

# Display order of karats within a group (ascending, matches the target report).
_KARAT_ORDER = ["14KT", "18KT", "22KT", "24KT"]
_BANDS = {"24KT", "22KT", "18KT", "14KT"}


def karat_from_melting(melting: float) -> str | None:
    if 99.0 <= melting <= 100.0:
        return "24KT"
    if 91.0 <= melting <= 92.5:
        return "22KT"
    if 74.5 <= melting <= 76.0:
        return "18KT"
    if 57.5 <= melting <= 59.8:
        return "14KT"
    return None


def karat_from_variant(variant, other=None) -> str | None:
    """Karat from a Variant Name.

    Uses a decimal melting % if present (e.g. "G-NA-91.80-YG" -> 91.80), else a
    literal "NNKT" (e.g. "PG-NA-24KT-YG" -> 24KT). Returns ``other`` when a
    value is present but falls outside the karat bands; None when no melting /
    karat value is found at all.
    """
    s = str(variant) if variant is not None else ""
    m = re.search(r"(\d+\.\d+)", s)            # decimal melting %
    if m:
        return karat_from_melting(float(m.group(1))) or other
    k = re.search(r"(\d+)\s*KT", s, re.I)      # literal karat, e.g. 24KT
    if k:
        band = f"{int(k.group(1))}KT"
        return band if band in _BANDS else other
    return None


@dataclass
class GroupsalesResult:
    rows: list[dict]          # flat preview: {Groupsales, Karat, Net Wt, Pg Wt}
    grand_net: float
    grand_pg: float
    date: str
    skipped: int              # rows whose melting matched no karat band


# --- styling (matches the target file) -------------------------------------
_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEAD_FILL = PatternFill("solid", fgColor=Color(theme=0, tint=-0.249977111117893))
_TOTAL_FILL = PatternFill("solid", fgColor=Color(theme=9, tint=0.3999755851924192))
_NUM_FMT = "0.000"
_FONT = "Cambria"
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_COL_WIDTHS = {1: 41.6, 2: 16.7, 3: 15.8, 4: 15.4}


def _find_header(grid: list[list]):
    """Row index of the header (contains 'Groupsales') and {label: col}."""
    for r, row in enumerate(grid):
        labels = {str(v).strip(): c for c, v in enumerate(row)
                  if v not in (None, "")}
        if "Groupsales" in labels:
            return r, labels
    return None, {}


def _extract_date(grid: list[list]) -> str:
    """Use the To Date from the preamble, formatted DD.MM.YYYY."""
    text = "\n".join(str(v) for row in grid for v in row
                     if isinstance(v, str))
    date_re = r"([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})"
    m = re.search(r"To\s*Date\s*:?-?\s*" + date_re, text, re.I)
    if not m:
        m = re.search(r"From\s*Date\s*:?-?\s*" + date_re, text, re.I)
    return m.group(1).replace("/", ".") if m else ""


def _aggregate(grid: list[list], labels: dict, header_row: int):
    need = ("Groupsales", "Variant Name", "Net Wt", "Pg Wt")
    missing = [c for c in need if c not in labels]
    if missing:
        raise LossReportError(
            "Groupsales report: missing column(s): " + ", ".join(missing))
    gi, vi = labels["Groupsales"], labels["Variant Name"]
    ni, pi = labels["Net Wt"], labels["Pg Wt"]
    fi = labels.get("Metal Fineness")          # optional, last-resort only

    def cell(row, i):
        return row[i] if (i is not None and i < len(row)) else None

    agg: "OrderedDict[str, dict]" = OrderedDict()
    skipped = 0
    for r in range(header_row + 1, len(grid)):
        row = grid[r]
        group = "" if cell(row, gi) is None else str(cell(row, gi)).strip()
        if group.lower().startswith("grand"):
            continue

        karat = karat_from_variant(cell(row, vi))
        if karat is None:
            fineness = to_number(cell(row, fi))
            if fineness is not None:
                melting = fineness * 100 if fineness <= 2 else fineness
                karat = karat_from_melting(melting)
        if karat is None:
            # No usable karat -> not a real data row (empty/spacer rows too).
            skipped += 1
            continue
        # Rows with a blank Groupsales are still counted, under "(blank)".
        key = group if group else "(blank)"
        net = to_number(cell(row, ni)) or 0.0
        pg = to_number(cell(row, pi)) or 0.0
        agg.setdefault(key, {})
        bucket = agg[key].setdefault(karat, [0.0, 0.0])
        bucket[0] += net
        bucket[1] += pg
    if not agg:
        raise LossReportError("Groupsales report: no usable data rows found.")
    return agg, skipped


def add_groupsales_sheet(wb: openpyxl.Workbook,
                         file_bytes: bytes) -> GroupsalesResult:
    """Append the 'Groupsales Reports' pivot sheet to ``wb``."""
    grid = _grid_from_bytes(file_bytes)
    header_row, labels = _find_header(grid)
    if header_row is None:
        raise LossReportError(
            "Could not find a header row containing 'Groupsales'. "
            "Please make sure you uploaded the Groupsales export."
        )
    agg, skipped = _aggregate(grid, labels, header_row)
    date = _extract_date(grid)

    ws = wb.create_sheet("Groupsales Reports")

    # --- Row 1 + 2: title and date ------------------------------------------
    ws.merge_cells("A1:D1")
    ws["A1"].value = "Groupsales Reports"
    ws["A1"].font = Font(name=_FONT, bold=True, size=14)
    ws["A1"].alignment = _CENTER
    ws.merge_cells("A2:D2")
    ws["A2"].value = f"Date: {date}"
    ws["A2"].font = Font(name=_FONT, bold=True, size=10)
    ws["A2"].alignment = _CENTER

    # --- Row 3: header ------------------------------------------------------
    for col, label in enumerate(("Groupsales", "Karat", "Net Wt", "Pg Wt"),
                                start=1):
        cell = ws.cell(row=3, column=col, value=label)
        cell.font = Font(name=_FONT, bold=True, size=9)
        cell.alignment = _CENTER
        cell.fill = _HEAD_FILL

    # --- Data (groups in data/first-appearance order, ascending karat) ------
    preview: list[dict] = []
    grand_net = grand_pg = 0.0
    out = 4
    for group in agg:
        karats = [k for k in _KARAT_ORDER if k in agg[group]]
        g_net = g_pg = 0.0
        for i, karat in enumerate(karats):
            net, pg = agg[group][karat]
            g_net += net
            g_pg += pg
            ws.cell(row=out, column=1,
                    value=group if i == 0 else None).font = (
                        Font(name=_FONT, bold=True, size=9))
            ws.cell(row=out, column=2, value=karat)
            ws.cell(row=out, column=3, value=net)
            ws.cell(row=out, column=4, value=pg)
            preview.append({"Groupsales": group if i == 0 else "",
                            "Karat": karat, "Net Wt": round(net, 3),
                            "Pg Wt": round(pg, 3)})
            out += 1
        # group total
        tcell = ws.cell(row=out, column=1, value=f"{group} Total")
        ws.cell(row=out, column=3, value=g_net)
        ws.cell(row=out, column=4, value=g_pg)
        for c in range(1, 5):
            ws.cell(row=out, column=c).fill = _TOTAL_FILL
            ws.cell(row=out, column=c).font = Font(name=_FONT, bold=True, size=9)
        preview.append({"Groupsales": f"{group} Total", "Karat": "",
                        "Net Wt": round(g_net, 3), "Pg Wt": round(g_pg, 3)})
        grand_net += g_net
        grand_pg += g_pg
        out += 1

    # --- Grand Total --------------------------------------------------------
    ws.cell(row=out, column=1, value="Grand Total")
    ws.cell(row=out, column=3, value=grand_net)
    ws.cell(row=out, column=4, value=grand_pg)
    for c in range(1, 5):
        ws.cell(row=out, column=c).fill = _TOTAL_FILL
        ws.cell(row=out, column=c).font = Font(name=_FONT, bold=True, size=9)

    # --- Uniform formatting / borders --------------------------------------
    for r in range(1, out + 1):
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            cell.border = _BORDER
            if cell.alignment.horizontal is None:
                cell.alignment = _CENTER
            if cell.font is None or cell.font.name != _FONT:
                if r >= 3:
                    cell.font = Font(name=_FONT, size=9,
                                     bold=bool(cell.font and cell.font.bold))
            if c in (3, 4) and r >= 4:
                cell.number_format = _NUM_FMT
    for col, width in _COL_WIDTHS.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    ws.freeze_panes = ws.cell(row=4, column=1)

    return GroupsalesResult(rows=preview, grand_net=round(grand_net, 3),
                            grand_pg=round(grand_pg, 3), date=date,
                            skipped=skipped)


def process(file_bytes: bytes) -> tuple[bytes, GroupsalesResult]:
    """Build a single-sheet Groupsales pivot workbook. Returns (bytes, result)."""
    from report_common import new_workbook, workbook_bytes
    wb = new_workbook()
    result = add_groupsales_sheet(wb, file_bytes)
    return workbook_bytes(wb), result
