from __future__ import annotations

import math
import re
from collections import OrderedDict, defaultdict
from dataclasses import dataclass, field

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.colors import Color

from groupsales import karat_from_variant
from loss_report import LossReportError, _grid_from_bytes, to_number

_FILL_FIELDS = {
    "created": "created by",
    "tdate": "trans date",
    "doc": "doc no",
    "batch": "batch no",
    "oper": "operation",
    "wc": "wc name",
}

_VALUE_FIELDS = [
    ("issue_g", "Issue Gms Wt"),
    ("issue_pg", "Issue Pg"),
    ("ret_g", "Return Gms Wt"),
    ("ret_pg", "Return Pg"),
    ("unut_g", "Unutilized Gms Wt"),
    ("unut_pg", "Unutilized Pg"),
    ("uscrap", "Unutilized Scrap"),
    ("uscrap_pg", "Unutilized Scrap Pg"),
    ("usample", "Unutilized Sample"),
    ("usample_pg", "Unutilized Sample Pg"),
    ("loss_g", "Loss Gms Wt"),
    ("loss_pg", "Loss Pg Wt"),
]

# Normalized header label -> value field key.
_VALUE_HEADERS = {
    "issuegmswt": "issue_g",
    "issuepg": "issue_pg",
    "returngmswt": "ret_g",
    "returnpg": "ret_pg",
    "unutilizedgmswt": "unut_g",
    "unutilizedpg": "unut_pg",
    "unutilizedscrap": "uscrap",
    "unutilizedscrappg": "uscrap_pg",
    "unutilizedsample": "usample",
    "unutilizedsamplepg": "usample_pg",
    "lossgmswt": "loss_g",
    "losspgwt": "loss_pg",
}

# Operations that are genuine filling work; anything else is a "double entry".
_FILLING_OPS = {"B-GOLD-FILING", "CAST-FILLING"}
_REP_PREFIX = "REP"

_REMARK_ORDER = ["FINISH", "DOUBLE ENTRY", "REP"]
_KARAT_ORDER = ["24KT", "22KT", "18KT", "14KT"]


@dataclass
class FillingLossResult:
    rows: list[dict]                    # flat preview of the pivot
    grand: dict                         # {value_key: total}
    title: str
    date_range: str
    remark_counts: dict = field(default_factory=dict)   # rows per remark
    skipped: int = 0                    # data rows with no resolvable karat


def _norm(value) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]", "", str(value).lower())


def _num(value) -> float:
    n = to_number(value)
    return 0.0 if n is None else n


def _roundup(x: float, digits: int = 3) -> float:
    """Excel ROUNDUP: round away from zero to ``digits`` decimals."""
    if x is None:
        return 0.0
    f = 10 ** digits
    r = math.ceil(round(abs(x) * f, 6)) / f
    return -r if x < 0 else r


def _find_header(grid: list[list]):
    """Row index of the header (has Batch No / Wc Name / Variant Name) and the
    label->column map for both descriptive and value columns."""
    for r, row in enumerate(grid):
        norm = {_norm(v): c for c, v in enumerate(row) if v not in (None, "")}
        if {"batchno", "wcname", "variantname"} <= set(norm):
            fill = {k: norm[lbl.replace(" ", "")] for k, lbl in
                    _FILL_FIELDS.items() if lbl.replace(" ", "") in norm}
            values = {k: norm[h] for h, k in _VALUE_HEADERS.items()
                      if h in norm}
            variant = norm["variantname"]
            return r, fill, values, variant
    return None, {}, {}, None


def _extract_meta(grid: list[list]) -> tuple[str, str]:
    """Report title (from WC Group) and the From/To date band, if present."""
    text = "\n".join(str(v) for row in grid for v in row
                     if isinstance(v, str))
    # Title: WC Group :- GOLD-FILLING  ->  FILLING
    title = "FILLING"
    # Match "WC Group :- GOLD-FILLING" specifically (not "WC Group Type :-").
    m = re.search(r"WC\s*Group\s*:\s*-?\s*([A-Za-z0-9\- ]+)", text, re.I)
    if m:
        grp = m.group(1).strip().upper()
        if grp:
            title = re.sub(r"^GOLD[\s\-]+", "", grp).strip() or grp
    # Date band: From Date :- 01/06/2026 ... To Date :- 30/06/2026
    date_re = r"([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})"
    mf = re.search(r"From\s*Date\s*:?-?\s*" + date_re, text, re.I)
    mt = re.search(r"To\s*Date\s*:?-?\s*" + date_re, text, re.I)
    dot = lambda m: m.group(1).replace("/", ".") if m else ""
    if mf and mt:
        date_range = f"{dot(mf)} to {dot(mt)}"
    else:
        date_range = dot(mt) or dot(mf)
    return title, date_range


def _is_blank_batch(batch) -> bool:
    """True for an empty cell or the literal 'NONE' placeholder batch."""
    return batch in (None, "") or str(batch).strip().upper() == "NONE"


def _build_rows(grid, header_row, fill, values, variant_col):
    """Parse the data rows, forward-filling the descriptive block, and derive
    Karat, Roundup, minifs, Roundup2 and the remark for each row."""

    def cell(row, c):
        return row[c] if (c is not None and c < len(row)) else None

    rows: list[dict] = []
    prev = {k: None for k in fill}
    skipped = 0
    for r in range(header_row + 1, len(grid)):
        src = grid[r]
        # Skip the trailing "Total" summary row and fully-blank rows.
        first = cell(src, fill.get("created"))
        if isinstance(first, str) and first.strip().lower() == "total":
            continue
        if all(cell(src, c) in (None, "") for c in range(len(src))):
            continue

        d = {}
        for k, c in fill.items():
            v = cell(src, c)
            if v in (None, ""):
                v = prev[k]
            prev[k] = v
            d[k] = v

        d["variant"] = cell(src, variant_col)
        d["karat"] = karat_from_variant(d["variant"])
        for k, _label in _VALUE_FIELDS:
            d[k] = _num(cell(src, values.get(k)))

        # A row with neither a karat nor any weight is not real data.
        if d["karat"] is None and all(d[k] == 0.0 for k, _ in _VALUE_FIELDS):
            continue
        if d["karat"] is None:
            skipped += 1
            d["karat"] = "(blank)"

        d["roundup"] = _roundup(d["ret_pg"], 3)
        rows.append(d)

    if not rows:
        raise LossReportError(
            "Found the header but no data rows to process.")

    groups: "defaultdict[object, list]" = defaultdict(list)
    for d in rows:
        groups[d["batch"]].append(d)
    for members in groups.values():
        lo = min(m["roundup"] for m in members)
        for m in members:
            m["minifs"] = lo
            m["roundup2"] = m["roundup"] - lo
            m["countif"] = len(members)

    _assign_remarks(rows)
    return rows, skipped


def _assign_remarks(rows: list[dict]) -> None:
    """Fill the ``remark`` column by running the reference steps 11-16 one after
    another, in order — exactly the sequence done by hand in Excel."""
    for d in rows:
        d["remark"] = None

    # Step 11: Wc Name = REP-GOLD-FILLING-WK  ->  REP
    for d in rows:
        if str(d["wc"] or "").strip().upper().startswith(_REP_PREFIX):
            d["remark"] = "REP"

    # Step 12: Operation other than B-GOLD-FILING / CAST-FILLING  ->  DOUBLE
    # ENTRY (leave the REP rows already set by step 11 untouched).
    for d in rows:
        if d["remark"] is None and \
                str(d["oper"] or "").strip().upper() not in _FILLING_OPS:
            d["remark"] = "DOUBLE ENTRY"

    # Step 13: Roundup2 = 0 (rows not already REP / double-entry)  ->  FINISH
    for d in rows:
        if d["remark"] is None and abs(d["roundup2"]) < 1e-9:
            d["remark"] = "FINISH"

    # Step 14: Roundup2 <> 0 (remaining rows)  ->  DOUBLE ENTRY
    for d in rows:
        if d["remark"] is None:
            d["remark"] = "DOUBLE ENTRY"

    # Step 15: a batch may keep only ONE FINISH; every other FINISH of the same
    # batch becomes DOUBLE ENTRY.
    _dedup_finish(rows)

    # Step 16: batch no. = NONE (or blank)  ->  DOUBLE ENTRY
    for d in rows:
        if d["remark"] != "REP" and _is_blank_batch(d["batch"]):
            d["remark"] = "DOUBLE ENTRY"


def _dedup_finish(rows: list[dict]) -> None:
    """Step 15: within a batch only ONE row may stay FINISH; the rest become
    DOUBLE ENTRY.

    A batch gets more than one FINISH only when several of its rows tie at the
    batch's minimum Roundup — so, as expected at this step, every such duplicate
    row has Roundup2 = 0. We keep the first tied row in sheet order and demote
    the rest. Because the tied rows all belong to the same batch, this never
    changes the grand total or the batch's own total; it can only move weight
    between the FINISH and DOUBLE ENTRY lines of that batch's karat.
    """
    kept: set = set()
    for d in rows:
        if d["remark"] != "FINISH":
            continue
        batch = d["batch"]
        if batch in kept:
            d["remark"] = "DOUBLE ENTRY"      # duplicate FINISH -> double entry
        else:
            kept.add(batch)


def _pivot(rows: list[dict]):
    """Aggregate the 12 weight columns by (remark, karat)."""
    agg: "OrderedDict[str, OrderedDict[str, dict]]" = OrderedDict()
    for d in rows:
        by_karat = agg.setdefault(d["remark"], OrderedDict())
        bucket = by_karat.setdefault(
            d["karat"], {k: 0.0 for k, _ in _VALUE_FIELDS})
        for k, _label in _VALUE_FIELDS:
            bucket[k] += d[k]
    return agg


_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEAD_FILL = PatternFill("solid", fgColor=Color(theme=0, tint=-0.249977111117893))
_TOTAL_FILL = PatternFill("solid", fgColor=Color(theme=9, tint=0.3999755851924192))
_GRAND_FILL = PatternFill("solid", fgColor="FFFFC000")
_NUM_FMT = "0.000"
_FONT = "Cambria"
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _karats_in_order(by_karat) -> list[str]:
    ordered = [k for k in _KARAT_ORDER if k in by_karat]
    ordered += [k for k in by_karat if k not in _KARAT_ORDER]
    return ordered


# Columns of the transparency "Working" sheet (steps 1-16), in order.
_WORK_COLS = [
    ("batch", "Batch No"), ("oper", "Operation"), ("wc", "Wc Name"),
    ("karat", "Karat"), ("variant", "Variant Name"),
    ("ret_pg", "Return Pg"), ("roundup", "Roundup"), ("minifs", "minifs"),
    ("roundup2", "Roundup2"), ("countif", "count if"), ("remark", "remark"),
] + list(_VALUE_FIELDS)


def _add_working_sheet(wb: openpyxl.Workbook, rows: list[dict]) -> None:
    """Write the full row-by-row detail after steps 1-16 so the output can be
    verified column-by-column against the source 'working file' subsheet."""
    ws = wb.create_sheet("Working (steps 1-16)")
    for c, (_key, label) in enumerate(_WORK_COLS, start=1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.font = Font(name=_FONT, bold=True, size=9)
        cell.alignment = _CENTER
        cell.fill = _HEAD_FILL
        cell.border = _BORDER
    for r, d in enumerate(rows, start=2):
        for c, (key, _label) in enumerate(_WORK_COLS, start=1):
            val = d.get(key)
            if isinstance(val, float):
                val = round(val, 4)
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name=_FONT, size=9)
            cell.border = _BORDER
            if key in ("roundup", "minifs", "roundup2") or \
                    key in dict(_VALUE_FIELDS):
                cell.number_format = _NUM_FMT
    ws.freeze_panes = ws.cell(row=2, column=1)
    for c in range(1, len(_WORK_COLS) + 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(c)].width = 12


def add_filling_loss_sheet(wb: openpyxl.Workbook,
                           file_bytes: bytes) -> FillingLossResult:
    """Append the 'Filling Loss & Recovery Report' pivot sheet to ``wb``."""
    grid = _grid_from_bytes(file_bytes)
    header_row, fill, values, variant_col = _find_header(grid)
    if header_row is None:
        raise LossReportError(
            "Could not find a header row with 'Batch No', 'Wc Name' and "
            "'Variant Name'. Please upload the GOLD-FILLING loss export."
        )
    missing = [lbl for key, lbl in _VALUE_FIELDS if key not in values]
    if missing:
        raise LossReportError(
            "These weight column(s) were not found: " + ", ".join(missing))

    title, date_range = _extract_meta(grid)
    rows, skipped = _build_rows(grid, header_row, fill, values, variant_col)
    agg = _pivot(rows)

    n_cols = 2 + len(_VALUE_FIELDS)          # remark + karat + 12 values
    last_col = openpyxl.utils.get_column_letter(n_cols)
    ws = wb.create_sheet("Filling Loss & Recovery Report")

    # --- Title + date band --------------------------------------------------
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"].value = f"{title} - Loss & Recovery Report"
    ws["A1"].font = Font(name=_FONT, bold=True, size=16)
    ws["A1"].alignment = _CENTER
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"].value = date_range
    ws["A2"].font = Font(name=_FONT, bold=True, size=11)
    ws["A2"].alignment = _CENTER

    # --- Header row ---------------------------------------------------------
    hr = 3
    headers = ["remark", "Karat"] + [lbl for _k, lbl in _VALUE_FIELDS]
    for c, label in enumerate(headers, start=1):
        cell = ws.cell(row=hr, column=c, value=label)
        cell.font = Font(name=_FONT, bold=True, size=9)
        cell.alignment = _CENTER
        cell.fill = _HEAD_FILL
        cell.border = _BORDER

    preview: list[dict] = []
    grand = {k: 0.0 for k, _ in _VALUE_FIELDS}
    remark_counts: dict = {}
    out = hr + 1

    def write_values(row, bucket, bold=False):
        for i, (k, _label) in enumerate(_VALUE_FIELDS, start=3):
            cell = ws.cell(row=row, column=i, value=round(bucket[k], 4))
            cell.font = Font(name=_FONT, bold=bold, size=9)
            cell.alignment = _CENTER
            cell.number_format = _NUM_FMT
            cell.border = _BORDER

    for remark in [r for r in _REMARK_ORDER if r in agg] + \
            [r for r in agg if r not in _REMARK_ORDER]:
        by_karat = agg[remark]
        group_total = {k: 0.0 for k, _ in _VALUE_FIELDS}
        karats = _karats_in_order(by_karat)
        remark_counts[remark] = sum(
            1 for d in rows if d["remark"] == remark)
        for i, karat in enumerate(karats):
            bucket = by_karat[karat]
            rcell = ws.cell(row=out, column=1,
                            value=remark if i == 0 else None)
            rcell.font = Font(name=_FONT, bold=True, size=9)
            rcell.alignment = _CENTER
            rcell.border = _BORDER
            kcell = ws.cell(row=out, column=2, value=karat)
            kcell.font = Font(name=_FONT, size=9)
            kcell.alignment = _CENTER
            kcell.border = _BORDER
            write_values(out, bucket)
            for k, _ in _VALUE_FIELDS:
                group_total[k] += bucket[k]
                grand[k] += bucket[k]
            preview.append({"remark": remark if i == 0 else "",
                            "Karat": karat,
                            **{lbl: round(bucket[k], 3)
                               for k, lbl in _VALUE_FIELDS}})
            out += 1

        # group total
        tcell = ws.cell(row=out, column=1, value=f"{remark} Total")
        tcell.font = Font(name=_FONT, bold=True, size=9)
        tcell.alignment = _CENTER
        tcell.border = _BORDER
        ws.cell(row=out, column=2).border = _BORDER
        write_values(out, group_total, bold=True)
        for c in range(1, n_cols + 1):
            ws.cell(row=out, column=c).fill = _TOTAL_FILL
        preview.append({"remark": f"{remark} Total", "Karat": "",
                        **{lbl: round(group_total[k], 3)
                           for k, lbl in _VALUE_FIELDS}})
        out += 1
        # blank spacer row between groups
        for c in range(1, n_cols + 1):
            ws.cell(row=out, column=c).border = _BORDER
        out += 1

    # --- Grand total --------------------------------------------------------
    gcell = ws.cell(row=out, column=1, value="Grand Total")
    gcell.font = Font(name=_FONT, bold=True, size=10)
    gcell.alignment = _CENTER
    gcell.border = _BORDER
    ws.cell(row=out, column=2).border = _BORDER
    write_values(out, grand, bold=True)
    for c in range(1, n_cols + 1):
        ws.cell(row=out, column=c).fill = _GRAND_FILL
    preview.append({"remark": "Grand Total", "Karat": "",
                    **{lbl: round(grand[k], 3) for k, lbl in _VALUE_FIELDS}})

    # --- Widths / freeze ----------------------------------------------------
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 9
    for c in range(3, n_cols + 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(c)].width = 13
    ws.freeze_panes = ws.cell(row=hr + 1, column=1)

    # Transparency sheet (steps 1-16), verifiable vs the source 'working file'.
    _add_working_sheet(wb, rows)

    return FillingLossResult(
        rows=preview, grand={k: round(grand[k], 3) for k, _ in _VALUE_FIELDS},
        title=title, date_range=date_range, remark_counts=remark_counts,
        skipped=skipped)


def process(file_bytes: bytes) -> tuple[bytes, FillingLossResult]:
    """Build a single-sheet Filling Loss workbook. Returns (bytes, result)."""
    from report_common import new_workbook, workbook_bytes
    wb = new_workbook()
    result = add_filling_loss_sheet(wb, file_bytes)
    return workbook_bytes(wb), result
