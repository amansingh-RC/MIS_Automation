import io

import openpyxl

from lot_rejection import add_lot_rejection_sheet, process
from report_common import new_workbook, workbook_bytes


def make_input() -> bytes:
    """Mimic the raw export: preamble, header at a 'Trans Date' row, totals."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B2"] = "SR"
    ws["B6"] = "Lot Rejection Report"
    ws["B9"] = "From Date :- 24/06/2026\nTo Date :- 25/06/2026\n"
    # header at row 12 (cols B..N), with merged-style gaps
    hdr = {2: "Trans Date", 3: "Order No", 4: "Group No", 5: "Style Name",
           6: "Karat", 7: "Wt", 8: "Operation Name", 10: "Wc Name",
           12: "User Name", 14: "Remark"}
    for c, label in hdr.items():
        ws.cell(row=12, column=c, value=label)
    # Trans Date as Excel serial 46197 -> 24/06/2026
    rows = [
        [46197.0, "", "SR-42018", "", "91.70", 15.17, "DIAMOND-CUTTING", "",
         "DIAMOND-CUTTING-WK", "", "KUMAR-UJJWAL", "", "dubara bnana hai"],
        [46197.0, "CNC-KADA-0406", "SR-42018", "SAMPLE-PCS", "", 0.0,
         "DIAMOND-CUTTING", "", "DIAMOND-CUTTING-WK", "", "KUMAR-UJJWAL", "",
         "dubara bnana hai"],
        [46197.0, "", "SR-42461", "", "91.70", 207.45, "ROLLING", "",
         "ROLLING-WK", "", "KUMAR-UJJWAL", "", "dubara bnana hai"],
    ]
    r = 13
    for row in rows:
        for i, val in enumerate(row):
            ws.cell(row=r, column=2 + i, value=val)
        r += 1
    ws.cell(row=r, column=2, value="Grand Total")
    ws.cell(row=r, column=7, value=222.62)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def main():
    out_bytes, result = process(make_input())

    assert result.date_from == "24/06/2026", result.date_from
    assert result.date_to == "25/06/2026", result.date_to
    assert len(result.rows) == 3, len(result.rows)
    assert abs(result.total_wt - 222.62) < 1e-9, result.total_wt
    print("OK  dates, row count, total Wt")

    wb = openpyxl.load_workbook(io.BytesIO(out_bytes))
    ws = wb.active
    assert ws.title == "Lot Rejection Report"
    assert ws["A1"].value == "Lot Rejection Report"
    assert ws["A2"].value == "From Date :- 24/06/2026 To Date :- 25/06/2026"
    headers = [ws.cell(row=3, column=c).value for c in range(1, 14)]
    assert headers[0] == "Trans Date" and headers[6] == "Operation Name"
    assert headers[8] == "Wc Name" and headers[10] == "User Name"
    assert headers[12] == "Remark"
    # Trans Date serial converted to dd/mm/yyyy
    assert ws.cell(row=4, column=1).value == "24/06/2026", ws.cell(row=4, column=1).value
    # merged Operation/Wc/User on header and data rows
    merged = {str(m) for m in ws.merged_cells.ranges}
    assert "G3:H3" in merged and "G4:H4" in merged
    assert "A1:M1" in merged and "A2:M2" in merged
    # Grand Total row recomputed
    gt_row = ws.max_row
    assert ws.cell(row=gt_row, column=1).value == "Grand Total"
    assert abs(ws.cell(row=gt_row, column=6).value - 222.62) < 1e-9
    print("OK  template structure, merges, Grand Total")

    # --- combined workbook: lot + a second sheet from another tool ----------
    from loss_report import add_loss_sheet  # noqa: F401  (import check only)
    wb2 = new_workbook()
    add_lot_rejection_sheet(wb2, make_input())
    names = openpyxl.load_workbook(io.BytesIO(workbook_bytes(wb2))).sheetnames
    assert names == ["Lot Rejection Report"], names
    print("OK  combined-workbook builder (single tool)")
    print("\nALL TESTS PASSED")


if __name__ == "__main__":
    main()
