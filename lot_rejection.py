from __future__ import annotations

import re
from dataclasses import dataclass

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from loss_report import LossReportError, _grid_from_bytes, to_number
from report_common import fmt_date


@dataclass
class LotRejectionResult:
    rows: list[dict]
    total_wt: float
    date_from: str
    date_to: str

_SINGLE = [
    ("Trans Date", 1), ("Order No", 2), ("Group No", 3), ("Style Name", 4),
    ("Karat", 5), ("Wt", 6), ("Remark", 13),
]
_MERGED = [
    ("Operation Name", 7, 8), ("Wc Name", 9, 10), ("User Name", 11, 12),
]
_LAST_COL = 13

_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_GRAY = PatternFill("solid", fgColor="FFD9D9D9")
_WT_FMT = '0.00;-0.00;""'

_COL_WIDTHS = {1: 11, 2: 20, 3: 10, 4: 28, 5: 7, 6: 8, 7: 9, 8: 9,
               9: 12, 10: 10, 11: 9, 12: 8, 13: 14}

_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)


def _find_header(grid: list[list]):
    """Row index containing 'Trans Date' and a {label: source_col} map."""
    for r, row in enumerate(grid):
        labels = {str(v).strip(): c for c, v in enumerate(row)
                  if v not in (None, "")}
        if "Trans Date" in labels:
            return r, labels
    return None, {}


def _extract_dates(grid: list[list]) -> tuple[str, str]:
    text = "\n".join(str(v) for row in grid for v in row
                     if isinstance(v, str))
    date_re = r"([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})"
    m_from = re.search(r"From\s*Date\s*:?-?\s*" + date_re, text, re.I)
    m_to = re.search(r"To\s*Date\s*:?-?\s*" + date_re, text, re.I)
    return (m_from.group(1) if m_from else "",
            m_to.group(1) if m_to else "")


def add_lot_rejection_sheet(wb: openpyxl.Workbook,
                            file_bytes: bytes) -> LotRejectionResult:
    """Append the formatted 'Lot Rejection Report' sheet to ``wb``."""
    grid = _grid_from_bytes(file_bytes)
    header_row, labels = _find_header(grid)
    if header_row is None:
        raise LossReportError(
            "Could not find a header row containing 'Trans Date'. "
            "Please make sure you uploaded the Lot Rejection Report."
        )

    date_from, date_to = _extract_dates(grid)
    ws = wb.create_sheet("Lot Rejection Report")
    last_letter = get_column_letter(_LAST_COL)

    # --- Row 1: title -------------------------------------------------------
    ws.merge_cells(f"A1:{last_letter}1")
    t = ws["A1"]
    t.value = "Lot Rejection Report"
    t.font = Font(bold=True, size=20)
    t.alignment = _CENTER

    # --- Row 2: date band ---------------------------------------------------
    ws.merge_cells(f"A2:{last_letter}2")
    d = ws["A2"]
    d.value = f"From Date :- {date_from} To Date :- {date_to}"
    d.font = Font(bold=True, size=14)
    d.alignment = _CENTER

    # --- Row 3: headers -----------------------------------------------------
    hr = 3
    for label, col in _SINGLE:
        cell = ws.cell(row=hr, column=col, value=label)
        cell.font = Font(bold=True)
        cell.alignment = _CENTER
        cell.fill = _GRAY
    for label, c1, c2 in _MERGED:
        ws.merge_cells(start_row=hr, start_column=c1, end_row=hr, end_column=c2)
        cell = ws.cell(row=hr, column=c1, value=label)
        cell.font = Font(bold=True)
        cell.alignment = _CENTER
        cell.fill = _GRAY
        ws.cell(row=hr, column=c2).fill = _GRAY

    # --- Data rows ----------------------------------------------------------
    def src(row, label):
        c = labels.get(label)
        return row[c] if (c is not None and c < len(row)) else None

    rows: list[dict] = []
    total_wt = 0.0
    out_row = hr + 1
    for r in range(header_row + 1, len(grid)):
        srow = grid[r]
        # Skip the source Grand Total row and fully-empty rows.
        if any(str(v).strip() == "Grand Total" for v in srow):
            continue
        values = {label: src(srow, label)
                  for label, *_ in (_SINGLE + _MERGED)}
        if all(v in (None, "") for v in values.values()):
            continue

        trans = fmt_date(values["Trans Date"])
        wt_num = to_number(values["Wt"]) or 0.0
        total_wt += wt_num

        # single columns
        ws.cell(row=out_row, column=1, value=trans).alignment = _LEFT
        ws.cell(row=out_row, column=2,
                value=_clean(values["Order No"])).alignment = _LEFT
        ws.cell(row=out_row, column=3,
                value=_clean(values["Group No"])).alignment = _LEFT
        ws.cell(row=out_row, column=4,
                value=_clean(values["Style Name"])).alignment = _LEFT
        ws.cell(row=out_row, column=5,
                value=_clean(values["Karat"])).alignment = _CENTER
        wt_cell = ws.cell(row=out_row, column=6, value=wt_num or None)
        wt_cell.number_format = _WT_FMT
        wt_cell.alignment = _RIGHT
        ws.cell(row=out_row, column=13,
                value=_clean(values["Remark"])).alignment = _LEFT

        # merged columns
        for label, c1, c2 in _MERGED:
            ws.merge_cells(start_row=out_row, start_column=c1,
                           end_row=out_row, end_column=c2)
            ws.cell(row=out_row, column=c1,
                    value=_clean(values[label])).alignment = _LEFT

        _border_row(ws, out_row)
        rows.append({
            "Trans Date": trans,
            "Order No": _clean(values["Order No"]),
            "Group No": _clean(values["Group No"]),
            "Style Name": _clean(values["Style Name"]),
            "Operation Name": _clean(values["Operation Name"]),
            "Wc Name": _clean(values["Wc Name"]),
            "Wt": wt_num,
        })
        out_row += 1

    if not rows:
        raise LossReportError("Found the header but no Lot Rejection data rows.")

    # --- Grand Total row ----------------------------------------------------
    gt = ws.cell(row=out_row, column=1, value="Grand Total")
    gt.font = Font(bold=True)
    wt_total = ws.cell(row=out_row, column=6, value=round(total_wt, 2))
    wt_total.font = Font(bold=True)
    wt_total.number_format = _WT_FMT
    wt_total.alignment = _RIGHT
    for col in range(1, _LAST_COL + 1):
        ws.cell(row=out_row, column=col).fill = _GRAY
    _border_row(ws, out_row)

    # --- Borders on title/date/header + dimensions --------------------------
    for row in (1, 2, hr):
        _border_row(ws, row)
    for col, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[hr].height = 22
    ws.freeze_panes = ws.cell(row=hr + 1, column=1)

    return LotRejectionResult(rows=rows, total_wt=round(total_wt, 2),
                              date_from=date_from, date_to=date_to)


def _clean(value) -> str:
    return "" if value is None else str(value).strip()


def _border_row(ws, row: int) -> None:
    for col in range(1, _LAST_COL + 1):
        ws.cell(row=row, column=col).border = _BORDER


def process(file_bytes: bytes) -> tuple[bytes, LotRejectionResult]:
    """Build a single-sheet Lot Rejection workbook. Returns (xlsx_bytes, result)."""
    from report_common import new_workbook, workbook_bytes
    wb = new_workbook()
    result = add_lot_rejection_sheet(wb, file_bytes)
    return workbook_bytes(wb), result
