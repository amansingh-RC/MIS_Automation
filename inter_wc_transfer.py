from __future__ import annotations

import re
from collections import OrderedDict, defaultdict
from dataclasses import dataclass

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter

from loss_report import LossReportError, _grid_from_bytes, to_number

# Step 1: only these Source Workers are kept (matched case/space/punct-insensitively).
_SOURCE_FILTER = [
    "CAM", "CAST GOLD BUFFING HOD", "CAST-MAGNETIC-HOD",
    "CENTRAL-OFFICE-1F-HOD", "ELECTROPLATING-HOD", "ENAMEL HOD",
    "GOLD METAL SETTING HOD", "GOLD-ASSEMBLY-HOD", "GOLD-BUFFING-HOD",
    "GOLD-CASTING-HOD", "GOLD-FILLING-HOD", "GOLD-REPAIR-HOD",
    "GOLD-SEPERATION-HOD", "GOLD-SHORT-HOD", "HALLMARK-HOD", "HAMMERING-HOD",
    "HAND-CUTTING-HOD", "MALABAR-FG-HOD", "MEDIA-POLISH-HOD",
    "PACKING & QA HOD", "PHOTOGRAPHY HOD", "REFINING-HOD", "WAX",
]

_SHEET_NAME = "INTER WC GROUP TRANSFER OUTWARD"


@dataclass
class InterWcResult:
    rows: list[dict]           # flat preview of the pivot
    grand_weight: float
    grand_pg: float
    date: str
    source_count: int
    kept_rows: int


# --- styling ---------------------------------------------------------------
_THIN = Side(style="thin", color="FF000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEAD_FILL = PatternFill("solid", fgColor=Color(theme=0, tint=-0.249977111117893))
_TOTAL_FILL = PatternFill("solid", fgColor=Color(theme=9, tint=0.3999755851924192))
_GRAND_FILL = PatternFill("solid", fgColor="FFFFC000")
_NUM_FMT = "0.000"
_FONT = "Cambria"
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center")


def _norm(value) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value).lower())


def _num(value) -> float:
    n = to_number(value)
    return 0.0 if n is None else n


def _find_header(grid: list[list]):
    """Header row (has Source Worker + Dest Worker + Weight) and {label: col}."""
    for r, row in enumerate(grid):
        labels = {str(v).strip(): c for c, v in enumerate(row)
                  if v not in (None, "")}
        if {"Source Worker", "Dest Worker", "Weight"} <= set(labels):
            return r, labels
    return None, {}


def _extract_date(grid: list[list]) -> str:
    """Report date from the preamble To/From date, formatted DD.MM.YYYY."""
    text = "\n".join(str(v) for row in grid for v in row
                     if isinstance(v, str))
    date_re = r"([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})"
    mf = re.search(r"From\s*Date\s*:?-?\s*" + date_re, text, re.I)
    mt = re.search(r"To\s*Date\s*:?-?\s*" + date_re, text, re.I)
    dot = lambda m: m.group(1).replace("/", ".") if m else ""
    d_from, d_to = dot(mf), dot(mt)
    if d_from and d_to and d_from != d_to:
        return f"{d_from} To {d_to}"
    return d_to or d_from


def add_inter_wc_transfer_sheet(wb: openpyxl.Workbook,
                                file_bytes: bytes) -> InterWcResult:
    """Append the 'INTER WC GROUP TRANSFER OUTWARD' pivot sheet to ``wb``.

    Filters the raw export to the chosen Source Workers, then pivots by
    Source Worker → Dest Worker, summing Weight and Pg Metal Weight.
    """
    grid = _grid_from_bytes(file_bytes)
    header_row, labels = _find_header(grid)
    if header_row is None:
        raise LossReportError(
            "Could not find a header row with 'Source Worker', 'Dest Worker' "
            "and 'Weight'. Please upload the Inter WC Group Transfer export.")
    for col in ("Source Worker", "Dest Worker", "Weight", "Pg Metal Weight"):
        if col not in labels:
            raise LossReportError(f"Inter WC Transfer: column '{col}' not found.")
    sc, dc = labels["Source Worker"], labels["Dest Worker"]
    wc, pc = labels["Weight"], labels["Pg Metal Weight"]

    def cell(row, i):
        return row[i] if i < len(row) else None

    # Canonical (filter-list) spelling per normalized source, so rows that
    # differ only in case/spacing merge — like an Excel pivot.
    canon = {_norm(x): x for x in _SOURCE_FILTER}
    # agg[source] = {norm_dest: [dest_display, weight, pg]}
    agg: "OrderedDict[str, OrderedDict]" = OrderedDict()
    kept = 0
    for r in range(header_row + 1, len(grid)):
        row = grid[r]
        source = "" if cell(row, sc) is None else str(cell(row, sc)).strip()
        if not source or source.lower().startswith(("grand", "total")):
            continue
        nsrc = _norm(source)
        if nsrc not in canon:
            continue
        src = canon[nsrc]
        dest = "" if cell(row, dc) is None else str(cell(row, dc)).strip()
        dests = agg.setdefault(src, OrderedDict())
        entry = dests.setdefault(_norm(dest), [dest, 0.0, 0.0])
        entry[1] += _num(cell(row, wc))
        entry[2] += _num(cell(row, pc))
        kept += 1
    if not agg:
        raise LossReportError(
            "Inter WC Transfer: no rows matched the Source Worker filter.")

    date = _extract_date(grid)
    ws = wb.create_sheet(_SHEET_NAME)

    ws.merge_cells("A1:D1")
    ws["A1"].value = _SHEET_NAME
    ws["A1"].font = Font(name=_FONT, bold=True, size=14)
    ws["A1"].alignment = _CENTER
    ws.merge_cells("A2:D2")
    ws["A2"].value = f"DATE : {date}"
    ws["A2"].font = Font(name=_FONT, bold=True, size=11)
    ws["A2"].alignment = _CENTER

    for c, label in enumerate(("Source Worker", "Dest Worker", "Weight",
                               "Pg Metal Weight"), start=1):
        h = ws.cell(row=3, column=c, value=label)
        h.font = Font(name=_FONT, bold=True, size=10)
        h.alignment = _CENTER
        h.fill = _HEAD_FILL

    preview: list[dict] = []
    grand_w = grand_p = 0.0
    out = 4
    # Source workers and their dest workers in alphabetical (pivot) order.
    for source in sorted(agg, key=lambda s: s.casefold()):
        entries = sorted(agg[source].values(), key=lambda e: e[0].casefold())
        s_w = s_p = 0.0
        for i, (dest, w, p) in enumerate(entries):
            s_w += w
            s_p += p
            ws.cell(row=out, column=1,
                    value=source if i == 0 else None).alignment = _LEFT
            ws.cell(row=out, column=2, value=dest).alignment = _LEFT
            ws.cell(row=out, column=3, value=round(w, 3))
            ws.cell(row=out, column=4, value=round(p, 3))
            preview.append({"Source Worker": source if i == 0 else "",
                            "Dest Worker": dest, "Weight": round(w, 3),
                            "Pg Metal Weight": round(p, 3)})
            out += 1
        tc = ws.cell(row=out, column=1, value=f"{source} Total")
        ws.cell(row=out, column=3, value=round(s_w, 3))
        ws.cell(row=out, column=4, value=round(s_p, 3))
        for c in range(1, 5):
            ws.cell(row=out, column=c).fill = _TOTAL_FILL
            ws.cell(row=out, column=c).font = Font(name=_FONT, bold=True, size=10)
        preview.append({"Source Worker": f"{source} Total", "Dest Worker": "",
                        "Weight": round(s_w, 3),
                        "Pg Metal Weight": round(s_p, 3)})
        grand_w += s_w
        grand_p += s_p
        out += 1

    ws.cell(row=out, column=1, value="Grand Total")
    ws.cell(row=out, column=3, value=round(grand_w, 3))
    ws.cell(row=out, column=4, value=round(grand_p, 3))
    for c in range(1, 5):
        ws.cell(row=out, column=c).fill = _GRAND_FILL
        ws.cell(row=out, column=c).font = Font(name=_FONT, bold=True, size=11)

    # borders + number formats
    for r in range(3, out + 1):
        for c in range(1, 5):
            cell_ = ws.cell(row=r, column=c)
            cell_.border = _BORDER
            if c in (3, 4) and r >= 4 and isinstance(cell_.value, (int, float)):
                cell_.number_format = _NUM_FMT
    for col, w in zip("ABCD", (30, 30, 16, 18)):
        ws.column_dimensions[get_column_letter(ord(col) - 64)].width = w
    ws.freeze_panes = ws.cell(row=4, column=1)

    return InterWcResult(
        rows=preview, grand_weight=round(grand_w, 3),
        grand_pg=round(grand_p, 3), date=date, source_count=len(agg),
        kept_rows=kept)


def process(file_bytes: bytes) -> tuple[bytes, InterWcResult]:
    """Build a single-sheet Inter WC Transfer workbook. Returns (bytes, result)."""
    from report_common import new_workbook, workbook_bytes
    wb = new_workbook()
    result = add_inter_wc_transfer_sheet(wb, file_bytes)
    return workbook_bytes(wb), result
