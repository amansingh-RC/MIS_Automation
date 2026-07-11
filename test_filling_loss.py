import io

import openpyxl

from filling_loss import (_dedup_finish, _roundup, add_gold_buffing_sheet,
                          process)
from report_common import new_workbook, workbook_bytes

# Column layout of the raw GOLD-FILLING export (1-based), matching the real
# file: descriptive block in A..H, weights spread across the rest.
_COLS = {
    "created": 2, "tdate": 3, "doc": 4, "batch": 5, "oper": 6, "wc": 7,
    "variant": 8, "issue_g": 10, "issue_pg": 12, "ret_g": 14, "ret_pg": 15,
    "unut_g": 18, "unut_pg": 19, "uscrap": 21, "uscrap_pg": 22,
    "usample": 24, "usample_pg": 25, "loss_g": 27, "loss_pg": 28,
}
_HEADERS = {
    2: "Created By", 3: "Trans Date", 4: "Doc No", 5: "Batch No",
    6: "Operation", 7: "Wc Name", 8: "Variant Name", 10: "Issue Gms Wt",
    12: "Issue Pg", 14: "Return Gms Wt", 15: "Return Pg",
    18: "Unutilized Gms Wt", 19: "Unutilized Pg", 21: "Unutilized Scrap",
    22: "Unutilized Scrap Pg", 24: "Unutilized Sample",
    25: "Unutilized Sample Pg", 27: "Loss Gms Wt", 28: "Loss Pg Wt",
}


def _row(ws, r, **vals):
    for key, col in _COLS.items():
        if key in vals and vals[key] is not None:
            ws.cell(row=r, column=col, value=vals[key])


def make_input() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B6"] = "OPERATION WISE-LOSS & GAIN NEW"
    ws["B9"] = ("WC Group Type :- \nWC Group :- GOLD-FILLING\n"
                "From Date :- 01/06/2026\nTo Date :- 30/06/2026\n")
    for c, label in _HEADERS.items():
        ws.cell(row=12, column=c, value=label)

    r = 13
    # FINISH: single-row batch, roundup2 = 0.
    _row(ws, r, created="A", batch="B-1", oper="B-GOLD-FILING",
         wc="GOLD-FILLING-WK", variant="G-NA-91.70-YG",
         issue_g=100.0, ret_pg=50.0, loss_pg=1.0); r += 1
    # DOUBLE ENTRY vs FINISH within one batch (B-2): the higher return pg gets
    # roundup2 > 0 (double entry); the min gets roundup2 = 0 (finish).
    _row(ws, r, created="A", batch="B-2", oper="B-GOLD-FILING",
         wc="GOLD-FILLING-WK", variant="G-NA-75.00-YG",
         issue_g=40.0, ret_pg=30.0, loss_pg=2.0); r += 1
    _row(ws, r, created="A", batch="B-2", oper="B-GOLD-FILING",
         wc="GOLD-FILLING-WK", variant="G-NA-75.00-YG",
         issue_g=20.0, ret_pg=20.0, loss_pg=0.0); r += 1
    # Continuation row: descriptive block blank -> inherits B-2 above, so it
    # joins that batch. Its own return pg 20.0 ties the batch min -> FINISH,
    # but B-2 already has a FINISH, so dedup makes this DOUBLE ENTRY.
    _row(ws, r, issue_g=10.0, ret_pg=20.0, loss_pg=0.5,
         variant="G-NA-75.00-YG"); r += 1
    # REP: WC starts with REP -> always REP regardless of roundup2.
    _row(ws, r, created="A", batch="B-3", oper="B-GOLD-FILING",
         wc="REP-GOLD-FILLING-WK", variant="G-NA-91.70-YG",
         issue_g=5.0, ret_pg=4.0, loss_pg=0.1); r += 1
    # Non-filling operation -> DOUBLE ENTRY even with roundup2 = 0.
    _row(ws, r, created="A", batch="B-4", oper="POLISH",
         wc="GOLD-FILLING-WK", variant="G-NA-91.70-YG",
         issue_g=7.0, ret_pg=6.0, loss_pg=0.3); r += 1
    # NONE batch -> DOUBLE ENTRY.
    _row(ws, r, created="A", batch="NONE", oper="CAST-FILLING",
         wc="GOLD-FILLING-WK", variant="G-NA-75.00-YG",
         issue_g=8.0, ret_pg=0.0, loss_pg=0.0); r += 1
    # Trailing total row -> skipped.
    ws.cell(row=r, column=2, value="Total")
    ws.cell(row=r, column=10, value=999.0)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_roundup():
    assert _roundup(77.82579, 3) == 77.826
    assert _roundup(71.19, 3) == 71.19
    assert _roundup(0.4125, 3) == 0.413
    print("OK  roundup (away from zero, 3 dp)")


def test_dedup_finish_multiway():
    """Step 15 for the case that shows up in *other* sheets: a batch with
    several rows tied at the min Roundup (all FINISH) must end with exactly one
    FINISH; every other tied row becomes DOUBLE ENTRY. Other batches and
    already-DOUBLE-ENTRY rows are untouched."""
    # Batch B-T: 3-way tie (all FINISH candidates). Batch B-U: single FINISH.
    rows = [
        {"batch": "B-T", "remark": "FINISH"},
        {"batch": "B-T", "remark": "FINISH"},
        {"batch": "B-U", "remark": "FINISH"},
        {"batch": "B-T", "remark": "FINISH"},
        {"batch": "B-T", "remark": "DOUBLE ENTRY"},   # not a tie -> unchanged
        {"batch": "B-U", "remark": "REP"},            # non-finish -> unchanged
    ]
    _dedup_finish(rows)
    remarks = [r["remark"] for r in rows]
    # First B-T FINISH kept; the two later B-T FINISH rows demoted.
    assert remarks == ["FINISH", "DOUBLE ENTRY", "FINISH", "DOUBLE ENTRY",
                       "DOUBLE ENTRY", "REP"], remarks
    # Exactly one FINISH per batch.
    fin_by_batch = {}
    for r in rows:
        if r["remark"] == "FINISH":
            fin_by_batch[r["batch"]] = fin_by_batch.get(r["batch"], 0) + 1
    assert all(v == 1 for v in fin_by_batch.values()), fin_by_batch
    print("OK  step 15 dedup — multi-way tie keeps exactly one FINISH")


def make_buffing_input() -> bytes:
    """Buffing export: same columns as filling, but buffing operations, a
    'REPAIR-...' repair WC, and an ancillary operation that must become a
    double entry via step 12."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B6"] = "OPERATION WISE-LOSS & GAIN NEW"
    ws["B9"] = ("WC Group Type :- \nWC Group :- GOLD-BUFFING-HOD\n"
                "From Date :- 01/06/2026\nTo Date :- 30/06/2026\n")
    for c, label in _HEADERS.items():
        ws.cell(row=12, column=c, value=label)
    r = 13
    # Genuine buffing op, single-row batch -> FINISH
    _row(ws, r, created="A", batch="B-1", oper="B-GOLD-BUFFING",
         wc="GOLD-BUFFING-WK", variant="G-NA-91.70-YG",
         issue_g=100.0, ret_pg=50.0, loss_pg=1.0); r += 1
    # Second genuine buffing op (Buffing-2) -> FINISH
    _row(ws, r, created="A", batch="B-2", oper="Buffing-2",
         wc="GOLD-BUFFING-WK", variant="G-NA-75.00-YG",
         issue_g=40.0, ret_pg=30.0, loss_pg=2.0); r += 1
    # Ancillary operation (not buffing) -> DOUBLE ENTRY via step 12
    _row(ws, r, created="A", batch="B-3", oper="FOR EMPTY BATCH",
         wc="GOLD-BUFFING-WK", variant="G-NA-91.70-YG",
         issue_g=7.0, ret_pg=6.0, loss_pg=0.3); r += 1
    # Repair work centre -> REP
    _row(ws, r, created="A", batch="B-4", oper="B-GOLD-BUFFING",
         wc="REPAIR-GOLD-BUFFING-WK", variant="G-NA-91.70-YG",
         issue_g=5.0, ret_pg=4.0, loss_pg=0.1); r += 1
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_gold_buffing():
    wb = new_workbook()
    res = add_gold_buffing_sheet(wb, make_buffing_input())
    out = workbook_bytes(wb)
    assert res.title == "GOLD BUFFING", res.title
    # B-1, B-2 genuine -> FINISH; B-3 ancillary op -> DOUBLE ENTRY; B-4 -> REP
    assert res.remark_counts.get("FINISH") == 2, res.remark_counts
    assert res.remark_counts.get("DOUBLE ENTRY") == 1, res.remark_counts
    assert res.remark_counts.get("REP") == 1, res.remark_counts
    names = openpyxl.load_workbook(io.BytesIO(out)).sheetnames
    assert names == ["Gold Buffing Loss & Recovery",
                     "Gold Buffing Working"], names
    assert all(len(n) <= 31 for n in names), names
    print("OK  gold buffing — operation double-entry, REP, ≤31-char tabs")


def main():
    test_roundup()
    test_dedup_finish_multiway()
    test_gold_buffing()
    out_bytes, res = process(make_input())

    assert res.title == "FILLING", res.title
    assert res.date_range == "01.06.2026 to 30.06.2026", res.date_range
    print("OK  title + date band")

    # Remark counts: FINISH = B-1 + one row of B-2 = 2
    # DOUBLE ENTRY = B-2 higher row + continuation + POLISH + NONE = 4
    # REP = 1
    assert res.remark_counts.get("FINISH") == 2, res.remark_counts
    assert res.remark_counts.get("DOUBLE ENTRY") == 4, res.remark_counts
    assert res.remark_counts.get("REP") == 1, res.remark_counts
    print("OK  remark classification (REP / operation / NONE / dedup)")

    # Grand total issue = 100+40+20+10+5+7+8 = 190 (the 999 Total row skipped)
    assert abs(res.grand["issue_g"] - 190.0) < 1e-9, res.grand
    # Loss pg = 1+2+0+0.5+0.1+0.3+0 = 3.9
    assert abs(res.grand["loss_pg"] - 3.9) < 1e-9, res.grand
    print("OK  grand totals (total row skipped)")

    wb = openpyxl.load_workbook(io.BytesIO(out_bytes))
    assert wb.sheetnames == ["Filling Loss & Recovery Report",
                             "Filling Working (steps 1-16)"], wb.sheetnames
    ws = wb["Filling Loss & Recovery Report"]
    assert ws["A1"].value == "FILLING - Loss & Recovery Report"
    assert ws["A3"].value == "remark" and ws["B3"].value == "Karat"

    # FINISH group: 22KT (B-1, issue 100) and 18KT (B-2 min row, issue 20).
    def find(label):
        return next(r for r in range(4, ws.max_row + 1)
                    if ws.cell(row=r, column=1).value == label)

    fin = find("FINISH")
    karats = {ws.cell(row=fin + i, column=2).value for i in range(2)}
    assert karats == {"22KT", "18KT"}, karats
    ft = find("FINISH Total")
    assert abs(ws.cell(row=ft, column=3).value - 120.0) < 1e-9  # 100 + 20
    print("OK  pivot layout + FINISH total (dedup kept one per batch)")

    gt = find("Grand Total")
    assert abs(ws.cell(row=gt, column=3).value - 190.0) < 1e-9
    print("OK  grand total row")
    print("\nALL TESTS PASSED")


if __name__ == "__main__":
    main()
