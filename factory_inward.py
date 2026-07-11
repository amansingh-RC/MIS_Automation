from __future__ import annotations

import re
from collections import OrderedDict, defaultdict
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter

from groupsales import karat_from_variant
from loss_report import LossReportError, _grid_from_bytes, to_number

_POSITIONS_PATH = Path(__file__).parent / "assets" / "party_positions.xlsx"
# "OT" (Other) captures any karat value outside the four standard bands.
_KARAT_ORDER = ["24KT", "22KT", "18KT", "14KT", "OT"]


@dataclass
class FactoryInwardResult:
    rows: list[dict]
    grand_net: float
    grand_pg: float
    date_from: str
    date_to: str
    unlisted: list[str]        # parties not found in the position table
    skipped: int               # rows whose variant matched no karat band


def load_positions(file_bytes: bytes) -> "OrderedDict[str, int]":
    """Parse a position table into {name_lower: position} (min on duplicates)."""
    grid = _grid_from_bytes(file_bytes)
    header_row = None
    name_c = pos_c = None
    for r, row in enumerate(grid):
        labels = {str(v).strip().lower(): c for c, v in enumerate(row)
                  if v not in (None, "")}
        if "name" in labels and "position" in labels:
            header_row, name_c, pos_c = r, labels["name"], labels["position"]
            break
    if header_row is None:
        raise LossReportError("Position table needs 'Name' and 'Position' columns.")

    positions: "OrderedDict[str, int]" = OrderedDict()
    for row in grid[header_row + 1:]:
        name = row[name_c] if name_c < len(row) else None
        pos = to_number(row[pos_c]) if pos_c < len(row) else None
        if name is None or pos is None:
            continue
        key = str(name).strip().lower()
        if not key:
            continue
        pos = int(pos)
        if key not in positions or pos < positions[key]:
            positions[key] = pos
    return positions


def default_positions() -> "OrderedDict[str, int]":
    """Bundled position table; empty (no crash) if the file is missing."""
    try:
        return load_positions(_POSITIONS_PATH.read_bytes())
    except OSError:
        return OrderedDict()

_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEAD_FILL = PatternFill("solid", fgColor=Color(theme=0, tint=-0.249977111117893))
_TOTAL_FILL = PatternFill("solid", fgColor=Color(theme=9, tint=0.3999755851924192))
_BLOCK_FILL = PatternFill("solid", fgColor=Color(theme=4, tint=0.7999816888943144))
_NUM_FMT = "0.000"
_FONT = "Cambria"
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_COL_WIDTHS = {1: 8, 2: 47, 3: 12, 4: 14, 5: 16}
_LAST_COL = 5


def _find_header(grid: list[list]):
    for r, row in enumerate(grid):
        labels = {str(v).strip(): c for c, v in enumerate(row)
                  if v not in (None, "")}
        if "Party Name" in labels and "Variant Name" in labels:
            return r, labels
    return None, {}


def _extract_dates(grid: list[list]) -> tuple[str, str]:
    text = "\n".join(str(v) for row in grid for v in row
                     if isinstance(v, str))
    date_re = r"([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})"
    mf = re.search(r"From\s*Date\s*:?-?\s*" + date_re, text, re.I)
    mt = re.search(r"To\s*Date\s*:?-?\s*" + date_re, text, re.I)
    fmt = lambda m: m.group(1).replace("/", "-") if m else ""
    return fmt(mf), fmt(mt)


def _extract_trans_type(grid: list[list]) -> str:
    text = "\n".join(str(v) for row in grid for v in row
                     if isinstance(v, str))
    m = re.search(r"Trans[ \t]*Type[ \t]*:?-?[ \t]*([^\n]+)", text, re.I)
    return m.group(1).strip() if m else ""


def add_factory_inward_sheet(wb: openpyxl.Workbook, file_bytes: bytes,
                             positions: "OrderedDict[str, int]",
                             gen_date: str, kind: str = "Inward",
                             default_trans: str = "GOODS RECEIPT NOTE",
                             sheet_name: str = "Factory Inward"
                             ) -> FactoryInwardResult:
    """Append a Factory Inward/Outward pivot sheet to ``wb``.

    ``kind`` ("Inward"/"Outward") and the transaction type (read from the
    preamble, falling back to ``default_trans``) form the title; the logic is
    otherwise identical for both reports.
    """
    grid = _grid_from_bytes(file_bytes)
    header_row, labels = _find_header(grid)
    if header_row is None:
        raise LossReportError(
            "Could not find a header row with 'Party Name' and 'Variant Name'. "
            "Please upload the Factory Inward export."
        )
    for col in ("Party Name", "Variant Name", "Net Wt", "Pg Wt"):
        if col not in labels:
            raise LossReportError(f"Factory Inward: column '{col}' not found.")
    pc, vc = labels["Party Name"], labels["Variant Name"]
    nc, gc = labels["Net Wt"], labels["Pg Wt"]

    def cell(row, i):
        return row[i] if i < len(row) else None

    # --- aggregate: party -> karat -> [net, pg]
    agg: "OrderedDict[str, dict]" = OrderedDict()
    skipped = 0
    for r in range(header_row + 1, len(grid)):
        row = grid[r]
        party = "" if cell(row, pc) is None else str(cell(row, pc)).strip()
        if not party or party.lower().startswith("grand"):
            continue
        # Out-of-band karats are kept as "OT"; only rows with no karat value
        # at all are skipped.
        karat = karat_from_variant(cell(row, vc), other="OT")
        if karat is None:
            skipped += 1
            continue
        net = to_number(cell(row, nc)) or 0.0
        pg = to_number(cell(row, gc)) or 0.0
        bucket = agg.setdefault(party, {}).setdefault(karat, [0.0, 0.0])
        bucket[0] += net
        bucket[1] += pg
    if not agg:
        raise LossReportError("Factory Inward: no usable data rows found.")

    # --- order parties by position table, then number Sr. No. serially ------
    listed = [(positions[p.lower()], p) for p in agg if p.lower() in positions]
    listed.sort(key=lambda x: (x[0], x[1].casefold()))
    unlisted = sorted((p for p in agg if p.lower() not in positions),
                      key=str.casefold)
    ordered_parties = [p for _, p in listed] + unlisted
    ordered = [(i, p) for i, p in enumerate(ordered_parties, start=1)]

    date_from, date_to = _extract_dates(grid)
    trans_type = _extract_trans_type(grid) or default_trans
    ws = wb.create_sheet(sheet_name)
    last_letter = get_column_letter(_LAST_COL)

    ws.merge_cells(f"A1:{last_letter}1")
    ws["A1"].value = f"Factory {kind} ( {trans_type} ) Report Summary"
    ws["A1"].font = Font(name=_FONT, bold=True, size=14)
    ws["A1"].alignment = _CENTER
    ws.merge_cells("A2:C2")
    ws.merge_cells("D2:E2")
    ws["A2"].value = f"Date From - {date_from} To {date_to}"
    ws["D2"].value = f"Date: {gen_date}"
    for addr in ("A2", "D2"):
        ws[addr].font = Font(name=_FONT, bold=True, size=12)
        ws[addr].alignment = _CENTER

    for col, label in enumerate(("Sr. No.", "Party Name", "KARAT",
                                 "Net Wt", "Pg Wt"), start=1):
        c = ws.cell(row=3, column=col, value=label)
        c.font = Font(name=_FONT, bold=True, size=10)
        c.alignment = _CENTER
        c.fill = _HEAD_FILL

    # --- Data ---------------------------------------------------------------
    preview: list[dict] = []
    grand_net = grand_pg = 0.0
    karat_totals: dict[str, list[float]] = defaultdict(lambda: [0.0, 0.0])
    out = 4
    for sr_no, party in ordered:
        karats = [k for k in _KARAT_ORDER if k in agg[party]]
        p_net = p_pg = 0.0
        for i, karat in enumerate(karats):
            net, pg = agg[party][karat]
            p_net += net
            p_pg += pg
            karat_totals[karat][0] += net
            karat_totals[karat][1] += pg
            ws.cell(row=out, column=1, value=sr_no if i == 0 else None)
            ws.cell(row=out, column=2, value=party if i == 0 else None)
            ws.cell(row=out, column=3, value=karat)
            ws.cell(row=out, column=4, value=net)
            ws.cell(row=out, column=5, value=pg)
            preview.append({"Sr. No.": sr_no if i == 0 else "",
                            "Party Name": party if i == 0 else "",
                            "Karat": karat, "Net Wt": round(net, 3),
                            "Pg Wt": round(pg, 3)})
            out += 1
        ws.cell(row=out, column=2, value=f"{party} Total")
        ws.cell(row=out, column=4, value=p_net)
        ws.cell(row=out, column=5, value=p_pg)
        for c in range(1, _LAST_COL + 1):
            ws.cell(row=out, column=c).fill = _TOTAL_FILL
            ws.cell(row=out, column=c).font = Font(name=_FONT, bold=True, size=10)
        preview.append({"Sr. No.": "", "Party Name": f"{party} Total",
                        "Karat": "", "Net Wt": round(p_net, 3),
                        "Pg Wt": round(p_pg, 3)})
        grand_net += p_net
        grand_pg += p_pg
        out += 1

    # Grand total
    ws.cell(row=out, column=2, value="Grand Total")
    ws.cell(row=out, column=4, value=grand_net)
    ws.cell(row=out, column=5, value=grand_pg)
    for c in range(1, _LAST_COL + 1):
        ws.cell(row=out, column=c).fill = _TOTAL_FILL
        ws.cell(row=out, column=c).font = Font(name=_FONT, bold=True, size=10)
    main_last = out

    # --- Karat-only summary block (under KARAT/Net/Pg columns) --------------
    block = out + 2
    for c, label in ((3, "KARAT"), (4, "Net Wt"), (5, "Pg Wt")):
        cell_ = ws.cell(row=block, column=c, value=label)
        cell_.font = Font(name=_FONT, bold=True, size=10)
        cell_.alignment = _CENTER
        cell_.fill = _BLOCK_FILL
    br = block + 1
    for karat in [k for k in _KARAT_ORDER if k in karat_totals]:
        ws.cell(row=br, column=3, value=karat)
        ws.cell(row=br, column=4, value=karat_totals[karat][0])
        ws.cell(row=br, column=5, value=karat_totals[karat][1])
        br += 1
    ws.cell(row=br, column=3, value="Grand Total")
    ws.cell(row=br, column=4, value=grand_net)
    ws.cell(row=br, column=5, value=grand_pg)
    for c in range(3, _LAST_COL + 1):
        ws.cell(row=br, column=c).font = Font(name=_FONT, bold=True, size=10)
        ws.cell(row=br, column=c).fill = _TOTAL_FILL

    # --- borders + number formats -------------------------------------------
    for r in range(3, main_last + 1):
        for c in range(1, _LAST_COL + 1):
            cell_ = ws.cell(row=r, column=c)
            cell_.border = _BORDER
            if cell_.alignment.horizontal is None:
                cell_.alignment = _CENTER
            if cell_.font is None or cell_.font.name != _FONT:
                cell_.font = Font(name=_FONT, size=10)
            if c in (4, 5) and r >= 4:
                cell_.number_format = _NUM_FMT
    for r in range(block, br + 1):
        for c in range(3, _LAST_COL + 1):
            cell_ = ws.cell(row=r, column=c)
            cell_.border = _BORDER
            cell_.alignment = _CENTER
            if c in (4, 5) and r > block:
                cell_.number_format = _NUM_FMT

    for col, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = ws.cell(row=4, column=1)

    return FactoryInwardResult(
        rows=preview, grand_net=round(grand_net, 3), grand_pg=round(grand_pg, 3),
        date_from=date_from, date_to=date_to, unlisted=unlisted, skipped=skipped)


def process(file_bytes: bytes, positions=None,
            gen_date: str = "") -> tuple[bytes, FactoryInwardResult]:
    """Build a single-sheet Factory Inward workbook. Returns (bytes, result)."""
    from report_common import new_workbook, workbook_bytes
    if positions is None:
        positions = default_positions()
    wb = new_workbook()
    result = add_factory_inward_sheet(wb, file_bytes, positions, gen_date)
    return workbook_bytes(wb), result
