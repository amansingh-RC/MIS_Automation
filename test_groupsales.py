import io

import openpyxl

from groupsales import process, karat_from_melting


def test_karat_bands():
    assert karat_from_melting(99.99) == "24KT"
    assert karat_from_melting(91.8) == "22KT"
    assert karat_from_melting(75.0) == "18KT"
    assert karat_from_melting(58.5) == "14KT"
    assert karat_from_melting(80.0) is None
    print("OK  karat bands")


def make_input() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B6"] = "All Transaction Summary"
    ws["B9"] = "From Date :- 24/06/2026\nTo Date :- 24/06/2026\n"
    # header at row 12 (cols B..), with the columns we need
    hdr = {2: "Trans Date", 11: "Metal Fineness", 14: "Variant Name",
           15: "Net Wt", 16: "Pg Wt", 23: "Groupsales"}
    for c, label in hdr.items():
        ws.cell(row=12, column=c, value=label)
    # group, fineness, variant, net, pg
    data = [
        ("CUBAN", 0.918, "", 100.0, 90.0),   # 22KT via fineness
        ("CUBAN", 0.75, "", 20.0, 15.0),     # 18KT
        ("ALPHA", 0.75, "", 10.0, 8.0),      # 18KT
        ("ALPHA", 0.80, "", 5.0, 4.0),       # out of band, no variant -> skipped
        ("ZETA", None, "PG-NA-24KT-YG", 30.0, 28.0),  # 24KT via Variant Name
        ("", 0.999, "", 40.0, 38.0),         # blank Groupsales -> "(blank)", 24KT
    ]
    r = 13
    for group, fin, var, net, pg in data:
        ws.cell(row=r, column=23, value=group)
        if fin is not None:
            ws.cell(row=r, column=11, value=fin)
        if var:
            ws.cell(row=r, column=14, value=var)
        ws.cell(row=r, column=15, value=net)
        ws.cell(row=r, column=16, value=pg)
        r += 1
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def main():
    test_karat_bands()
    out_bytes, result = process(make_input())

    assert result.date == "24.06.2026", result.date
    assert result.skipped == 1, result.skipped              # the 0.80 row
    # grand: net 100+20+10+30+40 = 200 ; pg 90+15+8+28+38 = 179
    assert abs(result.grand_net - 200.0) < 1e-9, result.grand_net
    assert abs(result.grand_pg - 179.0) < 1e-9, result.grand_pg
    print("OK  date, skipped, grand totals")

    wb = openpyxl.load_workbook(io.BytesIO(out_bytes))
    ws = wb.active
    assert ws.title == "Groupsales Reports"
    assert ws["A1"].value == "Groupsales Reports"
    assert ws["A2"].value == "Date: 24.06.2026"
    assert [ws.cell(row=3, column=c).value for c in range(1, 5)] == \
        ["Groupsales", "Karat", "Net Wt", "Pg Wt"]

    # CUBAN first (data / first-appearance order), karat ascending: 18KT then 22KT
    assert ws["A4"].value == "CUBAN"
    assert ws["B4"].value == "18KT" and abs(ws["C4"].value - 20.0) < 1e-9
    assert ws["B5"].value == "22KT" and abs(ws["C5"].value - 100.0) < 1e-9
    assert ws["A6"].value == "CUBAN Total"
    assert abs(ws["C6"].value - 120.0) < 1e-9

    # ZETA (24KT) came from the Variant Name even though Metal Fineness is blank
    zeta = next(r for r in range(4, ws.max_row + 1)
                if ws.cell(row=r, column=1).value == "ZETA")
    assert ws.cell(row=zeta, column=2).value == "24KT"
    assert abs(ws.cell(row=zeta, column=3).value - 30.0) < 1e-9

    # blank-Groupsales rows are kept under "(blank)"
    blank = next(r for r in range(4, ws.max_row + 1)
                 if ws.cell(row=r, column=1).value == "(blank)")
    assert ws.cell(row=blank, column=2).value == "24KT"
    assert abs(ws.cell(row=blank, column=3).value - 40.0) < 1e-9

    # last row grand total
    gt = ws.max_row
    assert ws.cell(row=gt, column=1).value == "Grand Total"
    assert abs(ws.cell(row=gt, column=3).value - 200.0) < 1e-9
    print("OK  layout, first-appearance order, ascending karat, (blank) group")

    # pivot only — no detail sheet
    assert wb.sheetnames == ["Groupsales Reports"], wb.sheetnames
    print("OK  single pivot sheet (no detail sheet)")
    print("\nALL TESTS PASSED")


if __name__ == "__main__":
    main()
