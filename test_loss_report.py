"""Self-test for the Loss Report template builder.

Builds a synthetic raw export (original layout: Sample BEFORE Scrap, a preamble
with dates, and a totals row), processes it, and asserts the output reproduces
the target "Loss Report Summary" template.
"""

import io

import openpyxl

from loss_report import process, to_number

# Raw export layout (matches the real .xls): header at row 13, Sample col F,
# Scrap col G, data starts row 14 with a totals row first.
RAW_HEADERS = [
    "Wc Name", "Issue Quantity Pg", "Process Quantity Pg",
    "Unutilized Quantity Pg", "Unutilized Quantity Sample Pg",
    "Unutilized Quantity Scrap Pg", "Loss Quantity Pg", "Gain Pg", "Bal Pg",
]
HEADER_ROW = 13
FIRST_COL = 2  # column B

# name, issue, process, unutil, sample, scrap, loss, gain, bal
DATA = [
    ("1FLOOR-ALLOYING-WK", 4677.92, 4663.72, 14.13, 0.0, 0.0, 0.066092, -2.2e-09, 0.0),
    ("BALL-CHAIN-WK", 6492.87, 6195.48, 291.74, 0.0, 5.126, 0.52269, 0.0, 0.0),
    ("CAST RHODIUM PLATING-WK", 1909.17, 1913.85, 0.0, 0.0, 0.0, 0.21, -4.89, 0.0),
]


def make_input() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B2"] = "SR"
    ws["B6"] = "Monthly Loss Report Summary"
    ws["B9"] = ("WC Group Type :- \nFrom Date :- 22/06/2026\n"
                "To Date :- 23/06/2026\nLot No :-\n")
    for c, h in enumerate(RAW_HEADERS, start=FIRST_COL):
        ws.cell(row=HEADER_ROW, column=c, value=h)
    # totals row (blank name) right under the header
    ws.cell(row=HEADER_ROW + 1, column=FIRST_COL + 1, value=11725.92)
    r = HEADER_ROW + 2
    for row in DATA:
        for c, val in enumerate(row, start=FIRST_COL):
            ws.cell(row=r, column=c, value=val)
        r += 1
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def main():
    out_bytes, report = process(make_input())

    # --- dates extracted from preamble --------------------------------------
    assert report.date_from == "22-06-2026", report.date_from
    assert report.date_to == "23-06-2026", report.date_to
    print("OK  dates extracted from preamble:", report.date_from, "->", report.date_to)

    # --- totals row dropped -------------------------------------------------
    assert len(report.rows) == 3, f"expected 3 data rows, got {len(report.rows)}"
    print("OK  totals row dropped (3 data rows)")

    wb = openpyxl.load_workbook(io.BytesIO(out_bytes))
    ws = wb.active

    # --- template structure -------------------------------------------------
    assert ws["A1"].value == "Loss Report Summary"
    assert ws["A2"].value == "Date From : 22-06-2026 To 23-06-2026  "
    assert ws["H2"].value == "Date : 23-06-2026"
    headers = [ws.cell(row=3, column=c).value for c in range(1, 13)]
    expected_headers = [
        "Wc Name", "Issue Quantity Pg", "Process Quantity Pg",
        "Unutilized Quantity Pg", "Unutilized Quantity Scrap Pg",
        "Unutilized Quantity Sample Pg", "Loss Quantity Pg", "Gain Pg",
        "Bal Pg", None, "FINAL LOSS", "LOSS %",
    ]
    assert headers == expected_headers, headers
    print("OK  title / date band / header order (Scrap before Sample)")

    # --- column reorder applied to data (Scrap=E, Sample=F) -----------------
    # BALL-CHAIN had scrap=5.126, sample=0 in the raw export -> E=5.126, F=0
    assert abs(ws.cell(row=5, column=5).value - 5.126) < 1e-9   # E = Scrap
    assert ws.cell(row=5, column=6).value == 0                  # F = Sample
    print("OK  Scrap/Sample columns correctly reordered in data")

    # --- formulas written ---------------------------------------------------
    assert ws["K4"].value == "=G4+H4"
    assert ws["L4"].value == '=IFERROR(K4/(C4+D4+E4+F4),"")'
    print("OK  FINAL LOSS / LOSS % written as formulas")

    # --- computed preview values --------------------------------------------
    by = {r["Wc Name"]: r for r in report.rows}
    checks = {
        "1FLOOR-ALLOYING-WK": (0.066092 - 2.2e-09, 4663.72 + 14.13),
        "BALL-CHAIN-WK": (0.52269, 6195.48 + 291.74 + 5.126),
        "CAST RHODIUM PLATING-WK": (0.21 - 4.89, 1913.85),
    }
    for name, (exp_final, denom) in checks.items():
        rec = by[name]
        assert abs(rec["Final Loss"] - round(exp_final, 4)) < 1e-4, name
        assert abs(rec["Loss %"] - exp_final / denom) < 1e-9, name
        print(f"OK  {name:26} final={rec['Final Loss']:.4f} loss%={rec['Loss %']:.4%}")

    assert to_number("(4.89)") == -4.89
    assert to_number("1,234.50") == 1234.50
    assert to_number(None) is None
    print("OK  to_number parsing verified")

    _test_xls_path()
    print("\nALL TESTS PASSED")


def _test_xls_path():
    """Fabricate a legacy .xls and confirm it is read and processed."""
    try:
        import xlwt
    except ImportError:
        print("--  skipping .xls test (xlwt not installed for fabrication)")
        return
    wb = xlwt.Workbook()
    ws = wb.add_sheet("Sheet1")
    ws.write(8, 1, "From Date :- 01/06/2026\nTo Date :- 02/06/2026")
    for c, h in enumerate(RAW_HEADERS, start=FIRST_COL):
        ws.write(HEADER_ROW - 1, c - 1, h)
    for i, row in enumerate(DATA):
        for c, val in enumerate(row, start=FIRST_COL):
            if val is not None:
                ws.write(HEADER_ROW + i, c - 1, val)
    buf = io.BytesIO()
    wb.save(buf)
    out_bytes, report = process(buf.getvalue())
    assert report.date_from == "01-06-2026" and report.date_to == "02-06-2026"
    wb2 = openpyxl.load_workbook(io.BytesIO(out_bytes))
    assert wb2.active["A1"].value == "Loss Report Summary"
    print("OK  legacy .xls read + processed (dates 01-06-2026 -> 02-06-2026)")


if __name__ == "__main__":
    main()
