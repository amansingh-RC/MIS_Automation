import io

import openpyxl

from inter_wc_transfer import process

# Raw layout (1-based) matching the real export: a leading blank column, then
# the descriptive block with Weight / Pg Metal Weight / Dest / Source workers.
_HEADERS = {
    2: "Doc No", 3: "Trans Date", 4: "Item Group", 5: "Variant Name",
    6: "Weight", 7: "Pg Metal Weight", 8: "Line Remark", 9: "Karat",
    10: "Dest Worker", 11: "Source Worker",
}


def _make_input(rows) -> bytes:
    """rows: (source, dest, weight, pg)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B5"] = "SALE REPORT"
    ws["B8"] = ("Trans Type :- INTER WC GROUP TRANSFER OUTWARD\n"
                "From Date :- 13/07/2026\nTo Date :- 13/07/2026\n")
    for c, label in _HEADERS.items():
        ws.cell(row=11, column=c, value=label)
    r = 12
    for source, dest, w, p in rows:
        ws.cell(row=r, column=6, value=w)
        ws.cell(row=r, column=7, value=p)
        ws.cell(row=r, column=10, value=dest)
        ws.cell(row=r, column=11, value=source)
        r += 1
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def main():
    rows = [
        ("GOLD-CASTING-HOD", "WAX", 10.0, 8.0),
        ("GOLD-CASTING-HOD", "WAX", 5.0, 4.0),       # merges with above
        ("GOLD-CASTING-HOD", "GOLD-FILLING-HOD", 20.0, 18.0),
        ("CAM", "CAD", 1.0, 0.75),
        ("SOME-OTHER-HOD", "CAD", 99.0, 99.0),       # not in filter -> dropped
        ("gold-casting-hod", "WAX", 2.0, 1.0),       # lower-case -> matches, merges
    ]
    out_bytes, res = process(_make_input(rows))

    assert res.date == "13.07.2026", res.date
    # Kept: all except SOME-OTHER-HOD = 5 rows; sources = {GOLD-CASTING-HOD, CAM}
    assert res.kept_rows == 5, res.kept_rows
    assert res.source_count == 2, res.source_count
    # Grand: weight 10+5+20+1+2 = 38 ; pg 8+4+18+0.75+1 = 31.75
    assert abs(res.grand_weight - 38.0) < 1e-9, res.grand_weight
    assert abs(res.grand_pg - 31.75) < 1e-9, res.grand_pg
    print("OK  filter (case-insensitive) + grand totals")

    wb = openpyxl.load_workbook(io.BytesIO(out_bytes))
    assert wb.sheetnames == ["INTER WC GROUP TRANSFER OUTWARD"], wb.sheetnames
    assert len(wb.sheetnames[0]) <= 31
    ws = wb.active
    assert ws["A1"].value == "INTER WC GROUP TRANSFER OUTWARD"
    assert ws["A2"].value == "DATE : 13.07.2026"
    assert [ws.cell(row=3, column=c).value for c in range(1, 5)] == \
        ["Source Worker", "Dest Worker", "Weight", "Pg Metal Weight"]
    # CAM sorts before GOLD-CASTING-HOD (alphabetical pivot order)
    assert ws["A4"].value == "CAM", ws["A4"].value
    print("OK  layout + alphabetical source order")

    # GOLD-CASTING-HOD: WAX merged (10+5+2=17), GOLD-FILLING-HOD (20). Total 37.
    gc_total = next(r for r in range(4, ws.max_row + 1)
                    if ws.cell(row=r, column=1).value == "GOLD-CASTING-HOD Total")
    assert abs(ws.cell(row=gc_total, column=3).value - 37.0) < 1e-9
    wax = next(r for r in range(4, ws.max_row + 1)
               if ws.cell(row=r, column=2).value == "WAX"
               and ws.cell(row=r, column=1).value in ("GOLD-CASTING-HOD", None))
    assert abs(ws.cell(row=wax, column=3).value - 17.0) < 1e-9    # merged
    print("OK  (source, dest) merge + per-source total")

    gt = next(r for r in range(4, ws.max_row + 1)
              if ws.cell(row=r, column=1).value == "Grand Total")
    assert abs(ws.cell(row=gt, column=3).value - 38.0) < 1e-9
    print("OK  grand total row")
    print("\nALL TESTS PASSED")


if __name__ == "__main__":
    main()
