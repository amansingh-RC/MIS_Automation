from __future__ import annotations

import io
import re
from dataclasses import dataclass

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter


class LossReportError(Exception):
    """Raised when the uploaded sheet cannot be understood."""


@dataclass
class ParsedReport:
    rows: list[dict]          # {Wc Name, Final Loss, Loss %} per data row
    date_from: str
    date_to: str

def _looks_like_xls(file_bytes: bytes) -> bool:
    """Legacy .xls files are OLE2 documents (magic D0 CF 11 E0).

    .xlsx/.xlsm are zip archives (magic 'PK'). Detect by content so it works
    regardless of extension or the MIME type the browser reports.
    """
    return file_bytes[:4] == b"\xd0\xcf\x11\xe0"


def _grid_from_bytes(file_bytes: bytes) -> list[list]:
    """Return the first sheet as a list of rows (each a list of cell values)."""
    if _looks_like_xls(file_bytes):
        try:
            import xlrd
        except ImportError as exc:
            raise LossReportError(
                "Reading .xls files needs the 'xlrd' package "
                "(pip install xlrd)."
            ) from exc
        book = xlrd.open_workbook(file_contents=file_bytes)
        sh = book.sheet_by_index(0)
        return [[sh.cell_value(r, c) for c in range(sh.ncols)]
                for r in range(sh.nrows)]

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise LossReportError(
            "Could not open this file as an Excel workbook. "
            "Please upload a .xls or .xlsx Loss Report sheet."
        ) from exc
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _norm(value) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]", "", str(value).lower())


def _match_field(norm_header: str) -> str | None:
    """Map a normalized header to a logical field. Order matters."""
    n = norm_header
    if n == "wcname":
        return "wc_name"
    if "scrap" in n:
        return "scrap"
    if "sample" in n:
        return "sample"
    if "unutilized" in n:
        return "unutilized"
    if "gain" in n:
        return "gain"
    if "loss" in n and "quantity" in n:
        return "loss"
    if "process" in n:
        return "process"
    if "issue" in n:
        return "issue"
    if n == "balpg" or n.startswith("bal"):
        return "bal"
    return None


def _find_header(grid: list[list]):
    """Find the header row (0-based) and a {field: col_index} map."""
    for r, row in enumerate(grid):
        col_map: dict[str, int] = {}
        for c, value in enumerate(row):
            field = _match_field(_norm(value))
            if field and field not in col_map:
                col_map[field] = c
        if "wc_name" in col_map:
            return r, col_map
    return None, {}


def to_number(value):
    """Convert a cell value to float; "(x)" -> negative; blanks -> None."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if s in ("", "-", "--"):
        return None
    negative = s.startswith("(") and s.endswith(")")
    if negative:
        s = s[1:-1]
    s = s.replace(",", "").replace("%", "").strip()
    try:
        n = float(s)
    except ValueError:
        return None
    return -n if negative else n


def _extract_dates(grid: list[list]) -> tuple[str, str]:
    """Pull From/To dates out of the preamble text (DD/MM/YYYY -> DD-MM-YYYY)."""
    text = "\n".join(
        str(v) for row in grid for v in row if isinstance(v, str)
    )
    date_re = r"([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})"
    m_from = re.search(r"From\s*Date\s*:?-?\s*" + date_re, text, re.I)
    m_to = re.search(r"To\s*Date\s*:?-?\s*" + date_re, text, re.I)
    fmt = lambda m: m.group(1).replace("/", "-") if m else ""
    return fmt(m_from), fmt(m_to)

OUTPUT_COLUMNS = [
    ("wc_name", "Wc Name"),
    ("issue", "Issue Quantity Pg"),
    ("process", "Process Quantity Pg"),
    ("unutilized", "Unutilized Quantity Pg"),
    ("scrap", "Unutilized Quantity Scrap Pg"),
    ("sample", "Unutilized Quantity Sample Pg"),
    ("loss", "Loss Quantity Pg"),
    ("gain", "Gain Pg"),
    ("bal", "Bal Pg"),
]
_SPACER_COL = 10   # J
_FINAL_COL = 11    # K
_PCT_COL = 12      # L
_LAST_COL = 12

_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_GRAY = PatternFill("solid", fgColor=Color(theme=0, tint=-0.249977111117893))
_ORANGE = PatternFill("solid", fgColor="FFFFC000")
_YELLOW = PatternFill("solid", fgColor="FFFFFF00")

_DATA_FMT = '[$-10409]#,##0.00;\\(#,##0.00\\);""'
_FINAL_FMT = "0.00"
_PCT_FMT = "0.000%"
_FONT = "Cambria"

_COL_WIDTHS = {
    1: 22, 2: 16.3, 3: 12.7, 4: 13.1, 5: 10.4, 6: 9.7, 7: 9.1, 8: 7.6,
    9: 9.3, 10: 4.9, 11: 9.0, 12: 9.7,
}


def _center(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def process(file_bytes: bytes) -> tuple[bytes, ParsedReport]:
    """Build the formatted Loss Report. Returns (xlsx_bytes, ParsedReport)."""
    grid = _grid_from_bytes(file_bytes)
    header_row, col_map = _find_header(grid)
    if header_row is None:
        raise LossReportError(
            "Could not find a header row containing 'Wc Name'. "
            "Please make sure you uploaded the correct Loss Report sheet."
        )
    missing = [m for m in ("loss", "gain", "process") if m not in col_map]
    if missing:
        pretty = {"loss": "Loss Quantity Pg", "gain": "Gain Pg",
                  "process": "Process Quantity Pg"}
        raise LossReportError(
            "These columns needed for the calculation were not found: "
            + ", ".join(pretty[m] for m in missing)
        )

    date_from, date_to = _extract_dates(grid)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Loss Report Summary"
    last_letter = get_column_letter(_LAST_COL)

    ws.merge_cells(f"A1:{last_letter}1")
    t = ws["A1"]
    t.value = "Loss Report Summary"
    t.font = Font(name=_FONT, bold=True, size=20)
    t.alignment = _center()

    ws.merge_cells("A2:G2")
    ws.merge_cells("H2:L2")
    left = ws["A2"]
    left.value = f"Date From : {date_from} To {date_to}  "
    right = ws["H2"]
    right.value = f"Date : {date_to}"
    for cell in (left, right):
        cell.font = Font(name=_FONT, bold=True, size=14)
        cell.alignment = _center()

    hr = 3
    for idx, (_field, label) in enumerate(OUTPUT_COLUMNS, start=1):
        cell = ws.cell(row=hr, column=idx, value=label)
        cell.font = Font(name=_FONT, bold=True, size=9)
        cell.alignment = _center()
        cell.fill = _GRAY
        cell.border = _BORDER
    sp = ws.cell(row=hr, column=_SPACER_COL)
    sp.fill = _YELLOW
    sp.border = _BORDER
    for col, label in ((_FINAL_COL, "FINAL LOSS"), (_PCT_COL, "LOSS %")):
        cell = ws.cell(row=hr, column=col, value=label)
        cell.font = Font(name=_FONT, bold=True, size=9)
        cell.alignment = _center()
        cell.fill = _ORANGE
        cell.border = _BORDER

    preview: list[dict] = []
    out_row = hr + 1
    for r in range(header_row + 1, len(grid)):
        src = grid[r]

        def get(field):
            c = col_map.get(field)
            return src[c] if (c is not None and c < len(src)) else None

        name_val = get("wc_name")
        name = "" if name_val is None else str(name_val).strip()
        nums = {f: to_number(get(f)) for f in
                ("issue", "process", "unutilized", "scrap", "sample",
                 "loss", "gain", "bal")}

        if not name or all(v is None for v in nums.values()):
            continue

        # Column A: work-center name
        a = ws.cell(row=out_row, column=1, value=name)
        a.font = Font(name=_FONT, size=9)
        a.alignment = _center()
        a.border = _BORDER

        # Columns B..I: numeric values
        for idx, (field, _label) in enumerate(OUTPUT_COLUMNS[1:], start=2):
            cell = ws.cell(row=out_row, column=idx, value=nums.get(field))
            cell.font = Font(name=_FONT, size=9)
            cell.alignment = _center()
            cell.number_format = _DATA_FMT
            cell.border = _BORDER

        # J: yellow spacer
        sp = ws.cell(row=out_row, column=_SPACER_COL)
        sp.fill = _YELLOW
        sp.border = _BORDER

        # K: FINAL LOSS  = Loss + Gain
        fcell = ws.cell(row=out_row, column=_FINAL_COL,
                        value=f"=G{out_row}+H{out_row}")
        fcell.font = Font(name=_FONT, bold=True, size=9)
        fcell.alignment = _center()
        fcell.number_format = _FINAL_FMT
        fcell.border = _BORDER

        # L: LOSS % = FINAL LOSS / (Process + Unutilized + Scrap + Sample)
        pcell = ws.cell(
            row=out_row, column=_PCT_COL,
            value=f"=IFERROR(K{out_row}/(C{out_row}+D{out_row}"
                  f"+E{out_row}+F{out_row}),\"\")")
        pcell.font = Font(name=_FONT, bold=True, size=9)
        pcell.alignment = _center()
        pcell.number_format = _PCT_FMT
        pcell.border = _BORDER

        # Preview (computed in Python so the app can show real numbers)
        loss = nums["loss"] or 0.0
        gain = nums["gain"] or 0.0
        final_loss = loss + gain
        denom = sum(nums[f] or 0.0
                    for f in ("process", "unutilized", "scrap", "sample"))
        preview.append({
            "Wc Name": name,
            "Final Loss": round(final_loss, 4),
            "Loss %": (final_loss / denom) if denom else None,
        })
        out_row += 1

    if not preview:
        raise LossReportError("Found the header but no data rows to process.")

    for col, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 25.5
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[hr].height = 36
    ws.freeze_panes = ws.cell(row=hr + 1, column=1)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), ParsedReport(rows=preview, date_from=date_from,
                                           date_to=date_to)
