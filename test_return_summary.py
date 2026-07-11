import io
from collections import OrderedDict

import openpyxl

from return_summary import process

# Raw file layout (1-based), matching the real GRN / INV exports: a leading
# blank column, then the descriptive block and weights.
_HEADERS = {
    2: "Trans Date", 3: "Doc No", 4: "Party Name", 5: "Style No",
    6: "Variant Name", 7: "Net Wt", 8: "Wastage Perc", 9: "Pg Wt",
    10: "Pg Wastage Wt", 11: "Line Remark",
}


def _make_raw(rows, trans_type) -> bytes:
    """rows: list of (doc, party, variant, net, pg)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B6"] = "All Transaction Summary"
    ws["B9"] = (f"Trans Type :- {trans_type}\n"
                "From Date :- 01/06/2026\nTo Date :- 30/06/2026\n")
    for c, label in _HEADERS.items():
        ws.cell(row=12, column=c, value=label)
    r = 13
    for doc, party, variant, net, pg in rows:
        ws.cell(row=r, column=3, value=doc)
        ws.cell(row=r, column=4, value=party)
        ws.cell(row=r, column=6, value=variant)
        ws.cell(row=r, column=7, value=net)
        ws.cell(row=r, column=9, value=pg)
        r += 1
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _positions():
    return OrderedDict([("royal chain private limited", 1), ("alpha", 2)])


def main():
    # Inward (GRN): a 24KT metal receipt + a 22KT return.
    inward = _make_raw([
        ("SR-GRN-GRN/26-27/1", "Royal Chain Private Limited",
         "G-NA-99.90-YG", 100.0, 99.0),       # 24KT -> METAL (Receipt)
        ("SR-GRN-GRN/26-27/2", "Royal Chain Private Limited",
         "G-NA-91.70-YG", 20.0, 18.0),        # 22KT -> RETURN
        ("SR-GRN-GRN/26-27/3", "Alpha",
         "G-NA-75.00-YG", 5.0, 4.0),          # 18KT -> RETURN
    ], "GOODS RECEIPT NOTE")
    # Outward (INV): a 24KT metal issue + a 22KT issue.
    outward = _make_raw([
        ("SR-INV-AST/26-27/1", "Royal Chain Private Limited",
         "G-NA-99.90-YG", 2.0, 2.0),          # 24KT -> ISSUE-METAL
        ("SR-INV-AST/26-27/2", "Royal Chain Private Limited",
         "G-NA-91.70-YG", 200.0, 180.0),      # 22KT -> ISSUE
        ("SR-INV-AST/26-27/3", "Alpha",
         "G-NA-75.00-YG", 50.0, 40.0),        # 18KT -> ISSUE
    ], "INV")

    out_bytes, res = process(inward, outward, _positions())

    assert res.parties == 2, res.parties
    assert res.date_from == "01.06.2026" and res.date_to == "30.06.2026"
    # Grand totals per category
    assert res.grand["METAL"] == [100.0, 99.0], res.grand
    assert res.grand["ISSUE-METAL"] == [2.0, 2.0], res.grand
    assert res.grand["ISSUE"] == [250.0, 220.0], res.grand
    assert res.grand["RETURN"] == [25.0, 22.0], res.grand
    print("OK  remark classification (Metal / Issue Metal / Issue / Return)")

    wb = openpyxl.load_workbook(io.BytesIO(out_bytes))
    assert wb.sheetnames == ["RETURN % SUMMARY"], wb.sheetnames
    ws = wb.active
    assert ws["A1"].value == "RETURN % SUMMARY"
    assert ws["C3"].value == "RECEIPT METAL" and ws["E3"].value == "ISSUE METAL"
    assert ws["G3"].value == "ISSUE" and ws["I3"].value == "RETURN"
    assert ws["M4"].value == "NET WT RETURN %"
    print("OK  layout + category headers")

    # Royal Chain 22KT row: Issue 200/180, Return 20/18.
    def find_row(party_karat):
        for r in range(5, ws.max_row + 1):
            if (ws.cell(row=r, column=2).value == party_karat):
                return r
        raise AssertionError(party_karat)

    r22 = next(r for r in range(5, ws.max_row + 1)
               if ws.cell(row=r, column=2).value == "22KT")
    assert abs(ws.cell(row=r22, column=7).value - 200.0) < 1e-9    # Issue Net
    assert abs(ws.cell(row=r22, column=9).value - 20.0) < 1e-9     # Return Net
    # Net Goods Net = Issue - Return = 180
    assert abs(ws.cell(row=r22, column=11).value - 180.0) < 1e-9
    # NET WT RETURN % = Return/Issue = 20/200 = 0.1
    assert abs(ws.cell(row=r22, column=13).value - 0.1) < 1e-9
    assert ws.cell(row=r22, column=13).number_format == "0.00%"
    print("OK  Net Goods (Issue−Return) + Return % (Return÷Issue)")

    # Karat order within Royal Chain: 24KT-METAL before 22KT.
    metal_r = find_row("24KT-METAL")
    assert metal_r < r22, (metal_r, r22)
    # 24KT-METAL row has Receipt Metal + Issue Metal, no Issue/Return.
    assert abs(ws.cell(row=metal_r, column=3).value - 100.0) < 1e-9
    assert abs(ws.cell(row=metal_r, column=5).value - 2.0) < 1e-9
    assert ws.cell(row=metal_r, column=7).value is None
    print("OK  24KT-METAL karat + column partitioning")

    gt = next(r for r in range(5, ws.max_row + 1)
              if ws.cell(row=r, column=1).value == "Grand Total")
    assert abs(ws.cell(row=gt, column=7).value - 250.0) < 1e-9
    print("OK  grand total row")
    print("\nALL TESTS PASSED")


if __name__ == "__main__":
    main()
