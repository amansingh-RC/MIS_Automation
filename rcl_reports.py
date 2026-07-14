from __future__ import annotations

import io
import math
import re
from collections import OrderedDict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from loss_report import LossReportError, _looks_like_xls

def _grid_first_sheet(file_bytes: bytes) -> list[list]:
    """First sheet as a list of rows (matches JS wb.SheetNames[0])."""
    if _looks_like_xls(file_bytes):
        try:
            import xlrd
        except ImportError as exc:
            raise LossReportError(
                "Reading .xls files needs the 'xlrd' package "
                "(pip install xlrd)."
            ) from exc
        book = xlrd.open_workbook(file_contents=file_bytes)
        sh = book.sheet_by_index(0)
        return [[sh.cell_value(r, c) for c in range(sh.ncols)]
                for r in range(sh.nrows)]

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise LossReportError(
            "Could not open this file as an Excel workbook."
        ) from exc
    ws = wb.worksheets[0]
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _s(value) -> str:
    """String of a cell, trimmed ('' for None)."""
    if value is None:
        return ""
    return str(value).strip()


def _pf(value) -> float:
    """parseFloat()-like: number -> float; junk -> 0.0."""
    if value is None or isinstance(value, bool):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", "")
    m = re.match(r"[-+]?\d*\.?\d+", s)
    return float(m.group(0)) if m else 0.0


def _r3(x: float) -> float:
    """Round half-up to 3 decimals (matches JS Math.round(x*1000)/1000)."""
    if x >= 0:
        return math.floor(x * 1000 + 0.5) / 1000
    return -math.floor(-x * 1000 + 0.5) / 1000


def _find_header_row(grid: list[list], marker: str) -> int:
    for i, row in enumerate(grid):
        if any(_s(c) == marker for c in row):
            return i
    return -1


def process_scrap(grid: list[list]) -> dict:
    header_row = _find_header_row(grid, "Stock Status")
    if header_row == -1:
        raise LossReportError(
            "Scrap report: header row not found "
            "(no column named 'Stock Status')."
        )
    headers = [_s(h) for h in grid[header_row]]
    ix = {h: i for i, h in enumerate(headers)}
    for col in ("Wcgroup Name", "Wc Name", "Gross Weight",
                "Metal Weight", "Stock Status"):
        if col not in ix:
            raise LossReportError(f"Scrap report: column '{col}' not found.")

    wcg, wcn = ix["Wcgroup Name"], ix["Wc Name"]
    grossc, metalc, statusc = (ix["Gross Weight"], ix["Metal Weight"],
                               ix["Stock Status"])

    def cell(row, i):
        return row[i] if i < len(row) else None

    rows = []
    for r in grid[header_row + 1:]:
        status = _s(cell(r, statusc)).upper()
        if status in ("SCRAP", "HL-SCRAP"):
            rows.append({
                "wcgroup": _s(cell(r, wcg)),
                "wcname": _s(cell(r, wcn)),
                "gross": _pf(cell(r, grossc)),
                "metal": _pf(cell(r, metalc)),
            })
    if not rows:
        raise LossReportError("Scrap report: no SCRAP / HL-SCRAP rows found.")

    groups = _group_and_sum(rows)
    total_gross = sum(r["gross"] for r in rows)
    total_metal = sum(r["metal"] for r in rows)
    return {"groups": groups, "total_gross": total_gross,
            "total_metal": total_metal, "row_count": len(rows)}


_EXCLUDE_ITEMS = {
    "ALLOY", "BEADS (GMS)", "COLOR STONE(GMS)", "CZ(GMS)",
    "SYNTHETIC STONE (GMS)", "PEARL (GMS)", "PEARL",
}
_TOTAL_ROW_RE = re.compile(r"\b(total|grand total|subtotal|sum)\b", re.I)


def process_stock(grid: list[list]) -> dict:
    header_row = _find_header_row(grid, "Wcgroup Name")
    if header_row == -1:
        raise LossReportError(
            "Stock report: header row not found "
            "(no column named 'Wcgroup Name')."
        )
    headers = [_s(h) for h in grid[header_row]]
    ix = {h: i for i, h in enumerate(headers)}
    if "Item Group" not in ix:
        raise LossReportError("Stock report: 'Item Group' column not found.")

    WCG = ix.get("Wcgroup Name")
    WCN = ix.get("Wc Name")
    PARTY = ix.get("Party Name")
    ITEM = ix.get("Item Group")
    GROSS = ix.get("Gross Weight")
    METAL = ix.get("Metal Weight")
    VNAME = ix.get("Variantt Name")  # note the double 't' (matches source)

    def cell(row, i):
        if i is None or i >= len(row):
            return None
        return row[i]

    total_rows = 0
    filtered_rows = 0
    merged: "OrderedDict[str, dict]" = OrderedDict()

    for r in grid[header_row + 1:]:
        non_blank = sum(1 for c in r if _s(c) != "")
        if non_blank <= 1:
            continue

        first_cell = _s(cell(r, 0)) or _s(cell(r, 1)) or _s(cell(r, 2))
        if _TOTAL_ROW_RE.search(first_cell):
            continue
        wcg_raw = _s(cell(r, WCG))
        wcn_raw = _s(cell(r, WCN))
        if _TOTAL_ROW_RE.search(wcg_raw) or _TOTAL_ROW_RE.search(wcn_raw):
            continue

        total_rows += 1

        party = _s(cell(r, PARTY))
        wcgroup = wcg_raw or party
        wcname = wcn_raw or party
        if not wcgroup and not wcname:
            continue

        item_raw = _s(cell(r, ITEM))
        item_up = item_raw.upper()
        if not item_raw:
            continue
        if item_up in _EXCLUDE_ITEMS:
            continue
        if item_raw.lower().replace(" ", "") == "pearl(gms)":
            continue

        if item_up == "OTHER METAL":
            if _s(cell(r, VNAME)).upper() != "OM":
                continue

        filtered_rows += 1
        gross = _pf(cell(r, GROSS))
        metal = _pf(cell(r, METAL))

        key = wcgroup.upper() + "|||" + wcname.upper()
        if key not in merged:
            merged[key] = {"wcgroup": wcgroup, "wcname": wcname,
                           "gross": 0.0, "metal": 0.0}
        merged[key]["gross"] += gross
        merged[key]["metal"] += metal

    if filtered_rows == 0:
        raise LossReportError("Stock report: no rows match the filter criteria.")

    entries = sorted(merged.values(),
                     key=lambda e: (e["wcgroup"], e["wcname"]))
    groups: "OrderedDict[str, list]" = OrderedDict()
    for e in entries:
        groups.setdefault(e["wcgroup"], []).append(e)

    total_gross = sum(e["gross"] for e in entries)
    total_metal = sum(e["metal"] for e in entries)
    return {"groups": groups, "total_gross": total_gross,
            "total_metal": total_metal, "group_count": len(groups),
            "total_rows": total_rows, "filtered_rows": filtered_rows}


def _group_and_sum(rows: list[dict]) -> "OrderedDict[str, list]":
    """Group rows by (wcgroup, wcname), summing weights; sorted output."""
    merged: dict[str, dict] = {}
    for r in rows:
        key = r["wcgroup"] + "|||" + r["wcname"]
        if key not in merged:
            merged[key] = {"wcgroup": r["wcgroup"], "wcname": r["wcname"],
                           "gross": 0.0, "metal": 0.0}
        merged[key]["gross"] += r["gross"]
        merged[key]["metal"] += r["metal"]

    entries = sorted(merged.values(),
                     key=lambda e: (e["wcgroup"], e["wcname"]))
    groups: "OrderedDict[str, list]" = OrderedDict()
    for e in entries:
        groups.setdefault(e["wcgroup"], []).append(e)
    return groups


_NUM_FMT = "#,##0.000"
_TITLE_FONT = Font(bold=True, size=14)
_SUB_FONT = Font(bold=True, size=11)
_HEAD_FONT = Font(bold=True)
_HEAD_FILL = PatternFill("solid", fgColor="FFD9D9D9")
_TOTAL_FONT = Font(bold=True)
_RIGHT = Alignment(horizontal="right")


def _write_data_rows(ws, groups, total_gross, total_metal, first_data_row):
    """Write the group/total/grand-total block; number-format weight cols."""
    last_row = first_data_row - 1
    for grp_name, entries in groups.items():
        g_gross = g_metal = 0.0
        for i, e in enumerate(entries):
            g_gross += e["gross"]
            g_metal += e["metal"]
            ws.append([grp_name if i == 0 else "", e["wcname"],
                       _r3(e["gross"]), _r3(e["metal"])])
            last_row += 1
        ws.append([f"{grp_name} Total", "", _r3(g_gross), _r3(g_metal)])
        last_row += 1
        for c in (1, 2, 3, 4):
            ws.cell(row=last_row, column=c).font = _TOTAL_FONT

    ws.append(["Grand Total", "", _r3(total_gross), _r3(total_metal)])
    last_row += 1
    for c in (1, 2, 3, 4):
        ws.cell(row=last_row, column=c).font = _TOTAL_FONT
    ws.append(["TOTAL SUM OF GROSS WEIGHT", "", _r3(total_gross), ""])
    last_row += 1
    ws.cell(row=last_row, column=1).font = _TOTAL_FONT

    for r in range(first_data_row, last_row + 1):
        for col in (3, 4):
            cell = ws.cell(row=r, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = _NUM_FMT
                cell.alignment = _RIGHT


def add_scrap_stock_sheets(wb: openpyxl.Workbook, file_bytes: bytes,
                           today: str) -> dict:
    """Append both SCRAP REPORT and STOCK REPORT sheets to ``wb``; return summary."""
    grid = _grid_first_sheet(file_bytes)
    scrap = process_scrap(grid)
    stock = process_stock(grid)

    ws = wb.create_sheet("SCRAP REPORT")
    ws.append(["ROYAL CHAIN LIMITED"])
    ws.append(["SCRAP & HL-SCRAP REPORT"])
    ws.append([])
    ws.append(["Wcgroup Name", "Wc Name", "Sum of Gross Weight",
               "Sum of Metal Weight"])
    ws["A1"].font = _TITLE_FONT
    ws["A2"].font = _SUB_FONT
    for c in range(1, 5):
        h = ws.cell(row=4, column=c)
        h.font = _HEAD_FONT
        h.fill = _HEAD_FILL
    _write_data_rows(ws, scrap["groups"], scrap["total_gross"],
                     scrap["total_metal"], first_data_row=5)
    for col, w in zip("ABCD", (30, 30, 22, 22)):
        ws.column_dimensions[col].width = w

    ws2 = wb.create_sheet("STOCK REPORT")
    ws2.append(["ROYAL CHAIN LIMITED"])
    ws2.append(["Daily Stock Report — Group-wise Summary"])
    ws2.append([f"Generated: {today}  |  Filters: ALLOY, Beads, Color Stone, "
                "CZ, Synthetic Stone, Pearl removed  |  OTHER METAL → OM only"])
    ws2.append([])
    ws2.append(["Wcgroup Name", "Wc Name", "Sum of Gross Weight",
                "Sum of Metal Weight"])
    ws2["A1"].font = _TITLE_FONT
    ws2["A2"].font = _SUB_FONT
    for c in range(1, 5):
        h = ws2.cell(row=5, column=c)
        h.font = _HEAD_FONT
        h.fill = _HEAD_FILL
    _write_data_rows(ws2, stock["groups"], stock["total_gross"],
                     stock["total_metal"], first_data_row=6)
    for col, w in zip("ABCD", (34, 34, 24, 24)):
        ws2.column_dimensions[col].width = w

    return {
        "scrap_rows": scrap["row_count"],
        "scrap_gross": scrap["total_gross"],
        "scrap_metal": scrap["total_metal"],
        "stock_filtered": stock["filtered_rows"],
        "stock_groups": stock["group_count"],
        "stock_gross": stock["total_gross"],
        "stock_metal": stock["total_metal"],
    }


def process_scrap_and_stock(file_bytes: bytes, today: str) -> tuple[bytes, dict]:
    """Run both reports on one input; return (xlsx_bytes, summary)."""
    from report_common import new_workbook, workbook_bytes
    wb = new_workbook()
    summary = add_scrap_stock_sheets(wb, file_bytes, today)
    return workbook_bytes(wb), summary


# ===========================================================================
# "Manish Report" variant — Scrap / Stock limited to a chosen set of Wcgroup
# names, in the compact "Stock Report _ Date _ Time" layout with green totals.
# ===========================================================================
_GREEN_FILL = PatternFill("solid", fgColor="FFA9D08E")
_THIN = Side(style="thin", color="FF000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")


def _norm_wcg(value) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value).lower())


def _apply_wcg_filter(result: dict, allowed) -> dict:
    """Keep only the Wcgroup Names in ``allowed`` (case/space/punct-insensitive)
    and recompute the totals from the kept groups."""
    if not allowed:
        return result
    allow = {_norm_wcg(x) for x in allowed}
    groups = OrderedDict((g, e) for g, e in result["groups"].items()
                         if _norm_wcg(g) in allow)
    out = dict(result)
    out["groups"] = groups
    out["total_gross"] = sum(e["gross"] for es in groups.values() for e in es)
    out["total_metal"] = sum(e["metal"] for es in groups.values() for e in es)
    if "group_count" in out:
        out["group_count"] = len(groups)
    if "row_count" in out:
        out["row_count"] = sum(len(es) for es in groups.values())
    return out


def _write_manish_block(ws, groups, title: str) -> None:
    """Title + gray header + group rows with green '{group} Total' rows."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    t = ws.cell(row=1, column=1, value=title)
    t.font = Font(bold=True, size=14)
    t.alignment = _CENTER
    for c, label in enumerate(("Wcgroup Name", "Wc Name", "Gross Weight",
                               "Metal Weight"), start=1):
        h = ws.cell(row=2, column=c, value=label)
        h.font = _HEAD_FONT
        h.fill = _HEAD_FILL
        h.alignment = _CENTER

    r = 3
    for grp, entries in groups.items():
        g_gross = g_metal = 0.0
        for i, e in enumerate(entries):
            g_gross += e["gross"]
            g_metal += e["metal"]
            ws.cell(row=r, column=1, value=grp if i == 0 else "")
            ws.cell(row=r, column=2, value=e["wcname"])
            ws.cell(row=r, column=3, value=_r3(e["gross"]))
            ws.cell(row=r, column=4, value=_r3(e["metal"]))
            r += 1
        ws.cell(row=r, column=1, value=f"{grp} Total")
        ws.cell(row=r, column=2, value="")
        ws.cell(row=r, column=3, value=_r3(g_gross))
        ws.cell(row=r, column=4, value=_r3(g_metal))
        for c in (1, 2, 3, 4):
            cell = ws.cell(row=r, column=c)
            cell.font = _TOTAL_FONT
            cell.fill = _GREEN_FILL
        r += 1

    for rr in range(2, r):
        for c in range(1, 5):
            cell = ws.cell(row=rr, column=c)
            cell.border = _BORDER
            if c in (3, 4) and isinstance(cell.value, (int, float)):
                cell.number_format = _NUM_FMT
                cell.alignment = _RIGHT
    for col, w in zip("ABCD", (34, 34, 18, 18)):
        ws.column_dimensions[col].width = w


def add_scrap_stock_manish_sheets(wb: openpyxl.Workbook, file_bytes: bytes,
                                  title_date: str, title_time: str,
                                  scrap_wcg=None, stock_wcg=None) -> dict:
    """Append filtered 'Scrap Report' and 'Stock Report' sheets (Manish layout)."""
    grid = _grid_first_sheet(file_bytes)
    scrap = _apply_wcg_filter(process_scrap(grid), scrap_wcg)
    stock = _apply_wcg_filter(process_stock(grid), stock_wcg)

    ws = wb.create_sheet("Scrap Report")
    _write_manish_block(
        ws, scrap["groups"],
        f"Scrap Report _ Date : {title_date} _ Time : {title_time}")
    ws2 = wb.create_sheet("Stock Report")
    _write_manish_block(
        ws2, stock["groups"],
        f"Stock Report _ Date : {title_date} _ Time : {title_time}")

    return {
        "scrap_groups": len(scrap["groups"]),
        "scrap_gross": scrap["total_gross"],
        "stock_groups": len(stock["groups"]),
        "stock_gross": stock["total_gross"],
    }
