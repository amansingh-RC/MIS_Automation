import io

import openpyxl

from factory_inward import process, karat_from_variant
from collections import OrderedDict


def test_variant_karat():
    assert karat_from_variant("G-NA-75.00-RG") == "18KT"
    assert karat_from_variant("G-NA-91.80-YG") == "22KT"
    assert karat_from_variant("PG-NA-99.50-YG") == "24KT"
    assert karat_from_variant("PG-NA-24KT-YG") == "24KT"     # literal fallback
    assert karat_from_variant("G-NA-80.00-YG") is None       # out of band
    assert karat_from_variant("G-NA-80.00-YG", other="OT") == "OT"
    assert karat_from_variant("NO-KARAT-HERE", other="OT") is None  # no value
    print("OK  karat from variant (decimal % + literal NNKT + OT)")


def make_input() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B6"] = "All Transaction Summary"
    ws["B9"] = "From Date :- 24/06/2026\nTo Date :- 24/06/2026\n"
    hdr = {2: "Trans Date", 3: "Party Name", 4: "Style No", 5: "Variant Name",
           6: "Net Wt", 8: "Pg Wt", 10: "Line Remark"}
    for c, label in hdr.items():
        ws.cell(row=12, column=c, value=label)
    # party, variant, net, pg
    data = [
        ("Royal Chain Private Limited", "PG-NA-24KT-YG", 100.0, 95.0),
        ("Royal Chain Private Limited", "G-NA-91.80-YG", 20.0, 18.0),
        ("RC-REFINISHING", "G-NA-75.00-RG", 10.0, 7.0),
        ("RC-REFINISHING", "G-NA-80.00-YG", 3.0, 2.0),   # out of band -> OT
        ("UNKNOWN PARTY", "G-NA-91.80-YG", 5.0, 4.0),   # not in position table
    ]
    r = 13
    for party, var, net, pg in data:
        ws.cell(row=r, column=3, value=party)
        ws.cell(row=r, column=5, value=var)
        ws.cell(row=r, column=6, value=net)
        ws.cell(row=r, column=8, value=pg)
        r += 1
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def main():
    test_variant_karat()
    # explicit position table: Royal Chain=1, RC-REFINISHING=2
    positions = OrderedDict([("royal chain private limited", 1),
                             ("rc-refinishing", 2)])
    out_bytes, result = process(make_input(), positions=positions,
                                gen_date="25-06-2026")

    assert result.date_from == "24-06-2026", result.date_from
    assert result.skipped == 0, result.skipped        # OT row is kept, not skipped
    assert result.unlisted == ["UNKNOWN PARTY"], result.unlisted
    # grand: net 100+20+10+3+5 = 138 ; pg 95+18+7+2+4 = 126
    assert abs(result.grand_net - 138.0) < 1e-9, result.grand_net
    assert abs(result.grand_pg - 126.0) < 1e-9, result.grand_pg
    print("OK  dates, skipped, unlisted, grand totals")

    wb = openpyxl.load_workbook(io.BytesIO(out_bytes))
    ws = wb.active
    assert ws.title == "Factory Inward"
    assert ws["A1"].value == "Factory Inward ( GOODS RECEIPT NOTE ) Report Summary"
    assert ws["A2"].value == "Date From - 24-06-2026 To 24-06-2026"
    assert ws["D2"].value == "Date: 25-06-2026"
    assert [ws.cell(row=3, column=c).value for c in range(1, 6)] == \
        ["Sr. No.", "Party Name", "KARAT", "Net Wt", "Pg Wt"]

    # Royal Chain (Sr.No 1) first, karat order 24 then 22
    assert ws["A4"].value == 1 and ws["B4"].value == "Royal Chain Private Limited"
    assert ws["C4"].value == "24KT" and ws["C5"].value == "22KT"
    assert ws["B6"].value == "Royal Chain Private Limited Total"
    assert abs(ws["D6"].value - 120.0) < 1e-9
    # RC-REFINISHING Sr.No 2
    assert ws["A7"].value == 2 and ws["B7"].value == "RC-REFINISHING"
    # UNKNOWN PARTY unlisted -> placed after, numbered 3 (max listed 2 + 1)
    rows = {ws.cell(row=r, column=2).value: ws.cell(row=r, column=1).value
            for r in range(4, ws.max_row + 1)}
    assert rows.get("UNKNOWN PARTY") == 3, rows
    print("OK  position order, Sr. No., karat order, totals")

    # Grand total + karat summary block present
    flat = [tuple(ws.cell(row=r, column=c).value for c in range(1, 6))
            for r in range(4, ws.max_row + 1)]
    assert any(t[1] == "Grand Total" for t in flat)
    assert any(t[2] == "KARAT" for t in flat)         # block header
    # out-of-band karat kept as "OT"
    assert "OT" in {t[2] for t in flat if t[2]}, flat
    print("OK  grand total + karat summary block + OT")

    # pivot only — no detail sheet
    assert wb.sheetnames == ["Factory Inward"], wb.sheetnames
    print("OK  single pivot sheet (no detail sheet)")
    print("\nALL TESTS PASSED")


if __name__ == "__main__":
    main()
