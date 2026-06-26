"""Shared helpers for building workbooks that hold one or more report sheets.

Each report module exposes an ``add_*_sheet(wb, ...)`` function that appends its
sheet(s) to a given workbook. That lets us either:
  * build a single-report workbook (one tab's download), or
  * build one master workbook with every tab's sheets combined.
"""

from __future__ import annotations

import io
from datetime import date, datetime, timedelta

import openpyxl

# Excel's 1900 date system epoch (openpyxl/Excel treat 1900 as a leap year, so
# the usable epoch for serial->date is 1899-12-30).
_EXCEL_EPOCH = datetime(1899, 12, 30)


def new_workbook() -> openpyxl.Workbook:
    """A fresh workbook. Its default empty 'Sheet' is removed by workbook_bytes."""
    return openpyxl.Workbook()


def workbook_bytes(wb: openpyxl.Workbook) -> bytes:
    """Drop the leftover empty default sheet, then return the .xlsx bytes."""
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        default = wb["Sheet"]
        empty = (default.max_row <= 1 and default.max_column <= 1
                 and default["A1"].value is None)
        if empty:
            del wb["Sheet"]
    if wb.sheetnames:
        wb.active = 0
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def excel_serial_to_date(value: float) -> datetime:
    return _EXCEL_EPOCH + timedelta(days=float(value))


def fmt_date(value, pattern: str = "%d/%m/%Y") -> str:
    """Format a cell that may be a datetime, an Excel serial, or text."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime(pattern)
    if isinstance(value, date):
        return value.strftime(pattern)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        # Plausible Excel date serials (skip tiny numbers that aren't dates).
        if value > 59:
            return excel_serial_to_date(value).strftime(pattern)
        return str(value)
    return str(value).strip()
