"""Self-test for Factory Outward (reuses Factory Inward logic)."""

import io
from collections import OrderedDict

import openpyxl

from factory_outward import process


def make_input(trans_type: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B6"] = "All Transaction Summary"
    ws["B8"] = (f"Trans Type :- {trans_type}\n"
                "From Date :- 24/06/2026\nTo Date :- 24/06/2026\n")
    hdr = {2: "Trans Date", 3: "Party Name", 4: "Style No", 5: "Variant Name",
           6: "Net Wt", 8: "Pg Wt", 10: "Line Remark"}
    for c, label in hdr.items():
        ws.cell(row=12, column=c, value=label)
    data = [
        ("Royal Chain Private Limited", "PG-NA-24KT-YG", 100.0, 95.0),
        ("RC-REFINISHING", "G-NA-75.00-RG", 10.0, 7.0),
        ("UNKNOWN PARTY", "G-NA-91.80-YG", 5.0, 4.0),
    ]
    r = 13
    for party, var, net, pg in data:
        ws.cell(row=r, column=3, value=party)
        ws.cell(row=r, column=5, value=var)
        ws.cell(row=r, column=6, value=net)
        ws.cell(row=r, column=8, value=pg)
        r += 1
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def main():
    positions = OrderedDict([("royal chain private limited", 1),
                             ("rc-refinishing", 2)])

    # Title uses the transaction type read from the preamble.
    out, result = process(make_input("DELIVERY CHALLAN"), positions=positions,
                          gen_date="26-06-2026")
    ws = openpyxl.load_workbook(io.BytesIO(out)).active
    assert ws.title == "Factory Outward", ws.title
    assert ws["A1"].value == "Factory Outward ( DELIVERY CHALLAN ) Report Summary", \
        ws["A1"].value
    assert [ws.cell(row=3, column=c).value for c in range(1, 6)] == \
        ["Sr. No.", "Party Name", "KARAT", "Net Wt", "Pg Wt"]
    # serial Sr. No. in position order; unlisted last
    assert ws["A4"].value == 1 and ws["B4"].value == "Royal Chain Private Limited"
    srnos = {ws.cell(row=r, column=2).value: ws.cell(row=r, column=1).value
             for r in range(4, ws.max_row + 1)}
    assert srnos.get("Royal Chain Private Limited") == 1, srnos
    assert srnos.get("RC-REFINISHING") == 2, srnos
    assert srnos.get("UNKNOWN PARTY") == 3, srnos
    assert abs(result.grand_net - 115.0) < 1e-9, result.grand_net   # 100+10+5
    print("OK  outward: title from trans type, sheet name, serial Sr.No, totals")

    # Falls back to the default trans type when the preamble lacks it.
    out2, _ = process(make_input(""), positions=positions, gen_date="26-06-2026")
    ws2 = openpyxl.load_workbook(io.BytesIO(out2)).active
    assert ws2["A1"].value == "Factory Outward ( DELIVERY CHALLAN ) Report Summary"
    print("OK  outward: default trans type fallback")
    print("\nALL TESTS PASSED")


if __name__ == "__main__":
    main()
