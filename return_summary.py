from __future__ import annotations

import re
from collections import OrderedDict, defaultdict
from dataclasses import dataclass

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter

from factory_inward import default_positions, load_positions  # noqa: F401
from groupsales import karat_from_variant
from loss_report import LossReportError, _grid_from_bytes, to_number

# Karat display order (reference pivot orders them this way: OT, then the
# 24KT metal band, then descending karats).
_KARAT_ORDER = ["OT", "24KT-METAL", "22KT", "18KT", "14KT"]

# Output categories (final-report header, source remark) in column order.
_CATEGORIES = [
    ("RECEIPT METAL", "METAL"),        # inward (GRN) + 24KT
    ("ISSUE METAL", "ISSUE-METAL"),    # outward (INV) + 24KT
    ("ISSUE", "ISSUE"),                # outward (INV) + non-24KT
    ("RETURN", "RETURN"),              # inward (GRN) + non-24KT
]


@dataclass
class ReturnSummaryResult:
    rows: list[dict]                   # flat preview of the pivot
    grand: dict                        # {remark: [net, pg]} + net-goods totals
    date_from: str
    date_to: str
    parties: int
    unlisted: list[str]                # parties not in the position table


# --- styling (mirrors the Factory report) ----------------------------------
_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEAD_FILL = PatternFill("solid", fgColor=Color(theme=0, tint=-0.249977111117893))
_TOTAL_FILL = PatternFill("solid", fgColor=Color(theme=9, tint=0.3999755851924192))
_GRAND_FILL = PatternFill("solid", fgColor="FFFFC000")
_NUM_FMT = "0.000"
_PCT_FMT = "0.00%"
_FONT = "Cambria"
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

_N_COLS = 2 + len(_CATEGORIES) * 2 + 2 + 2   # party+karat + 4×(net,pg) + net-goods + 2%
_COL_WIDTHS = {1: 42, 2: 12}


def _find_header(grid: list[list]):
    """Row index of the header (has Party Name + Variant Name) and {label: col}."""
    for r, row in enumerate(grid):
        labels = {str(v).strip(): c for c, v in enumerate(row)
                  if v not in (None, "")}
        if "Party Name" in labels and "Variant Name" in labels:
            return r, labels
    return None, {}


def _extract_dates(grid: list[list]) -> tuple[str, str]:
    """From/To dates from the preamble, formatted DD.MM.YYYY."""
    text = "\n".join(str(v) for row in grid for v in row
                     if isinstance(v, str))
    date_re = r"([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})"
    mf = re.search(r"From\s*Date\s*:?-?\s*" + date_re, text, re.I)
    mt = re.search(r"To\s*Date\s*:?-?\s*" + date_re, text, re.I)
    dot = lambda m: m.group(1).replace("/", ".") if m else ""
    return dot(mf), dot(mt)


def _num(value) -> float:
    n = to_number(value)
    return 0.0 if n is None else n


def _parse_file(file_bytes: bytes, is_inward: bool, agg: dict,
                parties: "OrderedDict[str, None]") -> tuple[str, str]:
    """Parse one raw file (inward GRN / outward INV) into ``agg``.

    Applies the Factory Inward/Outward logic (Party Name, Karat from the
    Variant Name with out-of-band = 'OT') but without the pivot, then assigns
    the remark by file + karat and folds the row into
    ``agg[party][karat][remark] = [net, pg]``.
    """
    grid = _grid_from_bytes(file_bytes)
    header_row, labels = _find_header(grid)
    if header_row is None:
        raise LossReportError(
            "Could not find a header row with 'Party Name' and 'Variant Name'. "
            "Please upload the correct Factory "
            f"{'Inward (GRN)' if is_inward else 'Outward (INV)'} export.")
    for col in ("Party Name", "Variant Name", "Net Wt", "Pg Wt"):
        if col not in labels:
            raise LossReportError(f"Return Summary: column '{col}' not found.")
    pc, vc = labels["Party Name"], labels["Variant Name"]
    nc, gc = labels["Net Wt"], labels["Pg Wt"]

    def cell(row, i):
        return row[i] if i < len(row) else None

    for r in range(header_row + 1, len(grid)):
        row = grid[r]
        party = "" if cell(row, pc) is None else str(cell(row, pc)).strip()
        if not party or party.lower().startswith(("grand", "total")):
            continue
        karat = karat_from_variant(cell(row, vc), other="OT")
        if karat is None:
            continue
        if karat == "24KT":
            karat = "24KT-METAL"
            remark = "METAL" if is_inward else "ISSUE-METAL"
        else:
            remark = "RETURN" if is_inward else "ISSUE"
        net = _num(cell(row, nc))
        pg = _num(cell(row, gc))
        parties.setdefault(party, None)
        bucket = agg.setdefault(party, {}).setdefault(karat, {})
        pair = bucket.setdefault(remark, [0.0, 0.0])
        pair[0] += net
        pair[1] += pg

    return _extract_dates(grid)


def _order_parties(parties, positions):
    """Position-table order (like Factory Inward): listed by position then name,
    unlisted alphabetically at the end."""
    listed = [(positions[p.lower()], p) for p in parties
              if p.lower() in positions]
    listed.sort(key=lambda x: (x[0], x[1].casefold()))
    unlisted = sorted((p for p in parties if p.lower() not in positions),
                      key=str.casefold)
    return [p for _, p in listed] + unlisted, unlisted


def _karats_for(party_agg) -> list[str]:
    present = list(party_agg.keys())
    ordered = [k for k in _KARAT_ORDER if k in party_agg]
    ordered += sorted((k for k in present if k not in _KARAT_ORDER),
                      reverse=True)
    return ordered


def add_return_summary_sheet(wb: openpyxl.Workbook, inward_bytes: bytes,
                             outward_bytes: bytes,
                             positions: "OrderedDict[str, int]"
                             ) -> ReturnSummaryResult:
    """Append the 'RETURN % SUMMARY' sheet built from the Factory Inward (GRN)
    and Factory Outward (INV) exports."""
    agg: "OrderedDict[str, dict]" = OrderedDict()
    parties: "OrderedDict[str, None]" = OrderedDict()
    df_in, dt_in = _parse_file(inward_bytes, True, agg, parties)
    df_out, dt_out = _parse_file(outward_bytes, False, agg, parties)
    if not agg:
        raise LossReportError("Return Summary: no usable data rows found.")
    date_from = df_out or df_in
    date_to = dt_out or dt_in

    ordered_parties, unlisted = _order_parties(parties, positions)

    ws = wb.create_sheet("RETURN % SUMMARY")
    last_letter = get_column_letter(_N_COLS)

    # --- Title + date band --------------------------------------------------
    ws.merge_cells(f"A1:{last_letter}1")
    ws["A1"].value = "RETURN % SUMMARY"
    ws["A1"].font = Font(name=_FONT, bold=True, size=16)
    ws["A1"].alignment = _CENTER
    ws.merge_cells(f"A2:{last_letter}2")
    ws["A2"].value = f"Monthly Summary From : {date_from} To {date_to}"
    ws["A2"].font = Font(name=_FONT, bold=True, size=11)
    ws["A2"].alignment = _CENTER

    # --- Row 3: category headers (merged pairs) -----------------------------
    for i, (label, _remark) in enumerate(_CATEGORIES):
        c0 = 3 + i * 2
        ws.merge_cells(start_row=3, start_column=c0,
                       end_row=3, end_column=c0 + 1)
        cell = ws.cell(row=3, column=c0, value=label)
        cell.font = Font(name=_FONT, bold=True, size=10)
        cell.alignment = _CENTER
        cell.fill = _HEAD_FILL
    ng_col = 3 + len(_CATEGORIES) * 2               # 11
    ws.merge_cells(start_row=3, start_column=ng_col,
                   end_row=3, end_column=ng_col + 1)
    ng = ws.cell(row=3, column=ng_col,
                 value="Net Goods issue\n After Return Less")
    ng.font = Font(name=_FONT, bold=True, size=10)
    ng.alignment = _CENTER
    ng.fill = _HEAD_FILL

    # --- Row 4: column headers ----------------------------------------------
    ws.cell(row=4, column=1, value="Party Name")
    ws.cell(row=4, column=2, value="KARAT")
    for i in range(len(_CATEGORIES) + 1):           # 4 categories + net-goods
        ws.cell(row=4, column=3 + i * 2, value="Net Wt")
        ws.cell(row=4, column=4 + i * 2, value="Pg Wt")
    ws.cell(row=4, column=ng_col + 2, value="NET WT RETURN %")
    ws.cell(row=4, column=ng_col + 3, value="PG WT RETURN %")
    for c in range(1, _N_COLS + 1):
        cell = ws.cell(row=4, column=c)
        cell.font = Font(name=_FONT, bold=True, size=10)
        cell.alignment = _CENTER
        cell.fill = _HEAD_FILL

    pct_net_col, pct_pg_col = ng_col + 2, ng_col + 3

    def write_line(out, party, karat, sums, bold, fill):
        """Write one data / total line from ``sums`` (remark -> [net, pg])."""
        if party is not None:
            pc = ws.cell(row=out, column=1, value=party)
            pc.alignment = _LEFT
        ws.cell(row=out, column=2, value=karat).alignment = _CENTER
        for i, (_label, remark) in enumerate(_CATEGORIES):
            pair = sums.get(remark)
            n = ws.cell(row=out, column=3 + i * 2,
                        value=round(pair[0], 3) if pair else None)
            g = ws.cell(row=out, column=4 + i * 2,
                        value=round(pair[1], 3) if pair else None)
            n.number_format = g.number_format = _NUM_FMT
        issue = sums.get("ISSUE", [0.0, 0.0])
        ret = sums.get("RETURN", [0.0, 0.0])
        ng_net, ng_pg = issue[0] - ret[0], issue[1] - ret[1]
        nn = ws.cell(row=out, column=ng_col, value=round(ng_net, 3))
        np_ = ws.cell(row=out, column=ng_col + 1, value=round(ng_pg, 3))
        nn.number_format = np_.number_format = _NUM_FMT
        pn = ws.cell(row=out, column=pct_net_col,
                     value=(ret[0] / issue[0]) if issue[0] else 0)
        pp = ws.cell(row=out, column=pct_pg_col,
                     value=(ret[1] / issue[1]) if issue[1] else 0)
        pn.number_format = pp.number_format = _PCT_FMT
        for c in range(1, _N_COLS + 1):
            cell = ws.cell(row=out, column=c)
            cell.border = _BORDER
            if cell.font is None or cell.font.name != _FONT:
                cell.font = Font(name=_FONT, size=10, bold=bold)
            if cell.alignment.horizontal is None:
                cell.alignment = _CENTER
            if fill:
                cell.fill = fill

    # --- Data ---------------------------------------------------------------
    preview: list[dict] = []
    grand: dict = defaultdict(lambda: [0.0, 0.0])
    out = 5
    for party in ordered_parties:
        party_agg = agg[party]
        p_tot: dict = defaultdict(lambda: [0.0, 0.0])
        karats = _karats_for(party_agg)
        for i, karat in enumerate(karats):
            sums = party_agg[karat]
            write_line(out, party if i == 0 else None, karat, sums,
                       bold=False, fill=None)
            for remark, pair in sums.items():
                p_tot[remark][0] += pair[0]
                p_tot[remark][1] += pair[1]
                grand[remark][0] += pair[0]
                grand[remark][1] += pair[1]
            preview.append({"Party Name": party if i == 0 else "",
                            "KARAT": karat,
                            **{lbl: round(sums.get(rk, [0, 0])[0], 3)
                               for lbl, rk in _CATEGORIES}})
            out += 1
        write_line(out, f"{party} Total", None, p_tot, bold=True,
                   fill=_TOTAL_FILL)
        out += 1

    # --- Grand total --------------------------------------------------------
    write_line(out, "Grand Total", None, grand, bold=True, fill=_GRAND_FILL)

    # --- Widths / freeze ----------------------------------------------------
    for c in range(3, _N_COLS + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13
    for col, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[3].height = 30
    ws.freeze_panes = ws.cell(row=5, column=1)

    return ReturnSummaryResult(
        rows=preview,
        grand={k: [round(v[0], 3), round(v[1], 3)] for k, v in grand.items()},
        date_from=date_from, date_to=date_to,
        parties=len(ordered_parties), unlisted=unlisted)


def process(inward_bytes: bytes, outward_bytes: bytes,
            positions=None) -> tuple[bytes, ReturnSummaryResult]:
    """Build a single-sheet RETURN % SUMMARY workbook. Returns (bytes, result)."""
    from report_common import new_workbook, workbook_bytes
    if positions is None:
        positions = default_positions()
    wb = new_workbook()
    result = add_return_summary_sheet(wb, inward_bytes, outward_bytes, positions)
    return workbook_bytes(wb), result
